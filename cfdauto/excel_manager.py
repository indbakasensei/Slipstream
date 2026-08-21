"""Excel schedule handling.

Design decision: **the workbook itself is the run database.**  The Status
column is the state machine (PENDING → RUNNING → DONE/FAILED), which gives
resume-after-interruption for free — restart the tool and it simply picks the
first row that is not DONE/SKIP.  Every result is *also* mirrored to a JSON
file per case, so even if the workbook is locked by the user the data is never
lost.

openpyxl (not pandas) is used for writing so the user's formatting, extra
columns and hand-written formulas are preserved untouched.
"""

from __future__ import annotations

import csv
import logging
import os
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import ExcelConfig
from .exceptions import ConfigError, ExcelWriteError
from .models import (Experiment, CaseResult, MetricValue, STATUS_DONE,
                     STATUS_FAILED, STATUS_PENDING, STATUS_RUNNING, STATUS_SKIP)
from .study_io import StudyIO

log = logging.getLogger(__name__)

WB_PARAM_PREFIX = "WBP:"          # extra Workbench-parameter columns
_HEADER_FONT = Font(bold=True)

# --------------------------------------------------------------------------- #
# Phase 8C — the universal output schema. These are NOT domain metrics: they
# are the run bookkeeping written for every template, and the legacy derived
# column kept for External-Aerodynamics backward compatibility.
# --------------------------------------------------------------------------- #
# (column_map_attr, number_format) — header = getattr(cfg.columns, attr).
_BOOKKEEPING_COLUMNS = (
    ("iterations", None), ("converged", None), ("error", None),
    ("started", None), ("finished", None), ("duration", "0.00"),
    ("case_dir", None),
)
# Legacy derived columns computed from a result and written when the workbook
# carries them. Data-driven — a column is only created/written when the
# template declares every metric it derives from, never a template-id branch.
# (column_map_attr, result_accessor, number_format, depends_on_metric_names)
_DERIVED_COLUMNS = (
    ("fl_fd", "fl_over_fd", "0.0000", ("lift", "drag")),
)
# Display formats for the template's declared metrics, keyed by metric name
# (template data — a metric the workbook carries keeps its hand-made look).
_METRIC_FORMATS = {
    "cl": "0.00000", "cd": "0.00000", "l_over_d": "0.0000",
    "lift": "0.000", "drag": "0.000",
}
# Fallback accessors for quantities a result exposes as computed properties
# rather than stored metrics (CL/CD is never stored by the aero strategy).
_METRIC_ACCESSORS = {
    "l_over_d": "cl_over_cd", "lift": "lift_n", "drag": "drag_n",
}


def _as_float(value: object) -> Optional[float]:
    """Cell value → float; ``None`` for blank. A present-but-unreadable value
    is logged (never silently ignored) and treated as absent."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        log.warning("Unreadable numeric output cell value %r — treated as "
                    "empty.", value)
        return None


class ExcelManager:
    """Owns the workbook: column mapping, experiment parsing, result writing."""

    def __init__(self, cfg: ExcelConfig, study_io: Optional["StudyIO"] = None):
        self.cfg = cfg
        self.path = cfg.path()
        self.wb = load_workbook(self.path)
        if cfg.sheet not in self.wb.sheetnames:
            raise ConfigError(
                f"Sheet '{cfg.sheet}' not found in {self.path}. "
                f"Available sheets: {self.wb.sheetnames}"
            )
        self.ws: Worksheet = self.wb[cfg.sheet]
        self._col: Dict[str, int] = {}          # header text -> column index
        self._wbp_cols: Dict[str, int] = {}     # WB param name -> column index
        # Phase 5: the study-I/O layer resolves which columns carry the
        # study's input parameters and builds Experiments from them. Built
        # here (before column mapping) so the required-column check and the
        # reader both go through the same template-driven boundary.
        #
        # Capability 3: a project passes its own template-resolved StudyIO
        # (via ``StudyIO.for_config``); without one we fall back to the
        # registry default so every existing caller is unchanged.
        self._study_io = study_io or StudyIO.default(cfg.columns)
        self._map_columns()

    @classmethod
    def for_config(cls, cfg) -> "ExcelManager":
        """Build a project-template-aware manager from a full :class:`Config`
        — its ``excel`` section plus its template's study-I/O mapping."""
        return cls(cfg.excel, study_io=StudyIO.for_config(cfg))

    # ------------------------------------------------------------------ #
    # Column discovery
    # ------------------------------------------------------------------ #
    def _map_columns(self) -> None:
        hdr_row = self.cfg.header_row
        for cell in self.ws[hdr_row]:
            if cell.value is None:
                continue
            name = str(cell.value).strip()
            self._col[name] = cell.column
            if name.upper().startswith(WB_PARAM_PREFIX):
                self._wbp_cols[name[len(WB_PARAM_PREFIX):].strip()] = cell.column

        # Phase 5: the required input columns come from the study definition
        # (via StudyIO) rather than a hardcoded (aoa, velocity) pair — same
        # headers today, correct for any future template.
        for required in self._study_io.input_column_headers():
            if required not in self._col:
                raise ConfigError(
                    f"Required input column '{required}' not found in header row "
                    f"{hdr_row} of sheet '{self.cfg.sheet}'. Found: {sorted(self._col)}"
                )

        # Create any missing output columns at the right-hand edge (headers
        # only — never reorders or touches the user's existing layout). The
        # output schema is template-driven (Phase 8C): Status + the template's
        # declared metric columns + universal bookkeeping + supported derived
        # columns — reproducing the legacy ColumnMap layout exactly for
        # External Aerodynamics.
        created = []
        for name in self._output_headers():
            if name not in self._col:
                col_idx = self.ws.max_column + 1
                cell = self.ws.cell(row=hdr_row, column=col_idx, value=name)
                cell.font = _HEADER_FONT
                self.ws.column_dimensions[get_column_letter(col_idx)].width = max(10, len(name) + 2)
                self._col[name] = col_idx
                created.append(name)
        if created:
            log.info("Added missing result columns to the schedule: %s", ", ".join(created))
            self.save()

    def _cell(self, row: int, header: str):
        return self.ws.cell(row=row, column=self._col[header])

    # ------------------------------------------------------------------ #
    # Output schema (Phase 8C) — template-driven, resolved via StudyIO.
    # The template owns its metrics; this manager only consumes them.
    # ------------------------------------------------------------------ #
    def _template(self):
        """The :class:`~cfdauto.platform.SimulationTemplate` driving this
        workbook's output schema (resolved through StudyIO). Never None —
        falls back to the registry default so every existing caller is
        unchanged."""
        tpl = getattr(self._study_io.exp_def, "template", None)
        if tpl is not None:
            return tpl
        from .platform import get_default_template  # lazy: no cycle
        return get_default_template()

    def _output_column_headers(self) -> List[Tuple[str, str]]:
        """Ordered ``(metric_name, header)`` pairs for the template's declared
        output metrics — the generic output-column contract, resolved at the
        StudyIO boundary against the project's ColumnMap."""
        return self._study_io.output_metric_columns()

    def _output_headers(self) -> List[str]:
        """Ordered output headers this manager creates/writes: Status, the
        template's metric columns, any legacy derived column the template
        supports, then the universal bookkeeping columns. For External
        Aerodynamics this reproduces ``ColumnMap.output_names()`` exactly."""
        c = self.cfg.columns
        headers = [c.status]
        headers += [header for _, header in self._output_column_headers()]
        for attr, _, _, depends in _DERIVED_COLUMNS:
            header = getattr(c, attr)
            if header in headers:
                continue
            if all(self._template().metric(m) is not None for m in depends):
                headers.append(header)
        for attr, _ in _BOOKKEEPING_COLUMNS:
            header = getattr(c, attr)
            if header not in headers:
                headers.append(header)
        return headers

    # ------------------------------------------------------------------ #
    # Reading the schedule
    # ------------------------------------------------------------------ #
    def read_experiments(self) -> List[Experiment]:
        """Parse every schedule row that has all inputs filled in.

        Phase 5: which columns are the study's inputs, and how a row maps to
        an :class:`Experiment`, is owned by the template-driven
        :class:`~cfdauto.study_io.StudyIO`. This method still owns the
        openpyxl cell access; StudyIO owns the mapping.

        NOTE: value *validation* (e.g. velocity > 0) is deliberately NOT done
        here — the orchestrator validates and marks such rows FAILED in the
        workbook, which is far more visible to the user than a console
        warning.
        """
        io = self._study_io
        param_names = io.input_parameter_names()
        headers = io.input_column_headers()
        c = self.cfg.columns
        experiments: List[Experiment] = []
        for row in range(self.cfg.header_row + 1, self.ws.max_row + 1):
            input_values = {name: self._cell(row, header).value
                            for name, header in zip(param_names, headers)}
            wbp_values = {name: self.ws.cell(row=row, column=col).value
                          for name, col in self._wbp_cols.items()}
            status = str(self._cell(row, c.status).value or "").strip().upper()
            exp, warning = io.interpret_row(row, input_values, wbp_values, status)
            if warning:
                log.warning("%s", warning)
            if exp is not None:
                experiments.append(exp)
        return experiments

    def pending(self, retry_failed: bool, rerun_stale_running: bool) -> List[Experiment]:
        """Rows that still need to run, honouring resume semantics."""
        todo: List[Experiment] = []
        for exp in self.read_experiments():
            s = exp.status
            if s in ("", STATUS_PENDING):
                todo.append(exp)
            elif s == STATUS_RUNNING and rerun_stale_running:
                log.warning("Row %d was left RUNNING by a previous session — re-queuing.", exp.row)
                todo.append(exp)
            elif s == STATUS_FAILED and retry_failed:
                todo.append(exp)
            elif s not in (STATUS_DONE, STATUS_SKIP, STATUS_FAILED, STATUS_RUNNING):
                log.warning("Row %d has unknown status '%s' — treating as pending.", exp.row, s)
                todo.append(exp)
        return todo

    # ------------------------------------------------------------------ #
    # Writing results
    # ------------------------------------------------------------------ #
    def mark_running(self, exp: Experiment, case_dir: str) -> None:
        c = self.cfg.columns
        self._cell(exp.row, c.status).value = STATUS_RUNNING
        self._cell(exp.row, c.error).value = ""
        self._cell(exp.row, c.case_dir).value = case_dir
        self.save()

    def _metric_cell(self, res: CaseResult, metric_name: str
                     ) -> Tuple[Optional[float], Optional[str]]:
        """(value, number_format) for one template metric of a result of any
        shape — template-attached generic or template-less legacy. Resolution
        honours the stored metric first, then derived/legacy accessors."""
        mv = res.metric(metric_name)
        if mv is not None:
            value = mv.value
        else:
            value = getattr(res, _METRIC_ACCESSORS.get(metric_name, ""), None)
        fmt = _METRIC_FORMATS.get(metric_name)
        if metric_name == "l_over_d" and value is not None:
            value = round(value, 4)
        return value, fmt

    def write_result(self, exp: Experiment, res: CaseResult, status: str) -> None:
        """Write one complete result row.  Numeric values (not formulas) are
        written so the sheet is immediately valid for pandas / resume logic.

        Phase 8C: the output columns are template-driven — the template's
        declared metrics (via StudyIO → ColumnMap), universal bookkeeping, and
        any supported legacy derived column. One generic writer produces an
        External Aerodynamics ``CL``/``CD``/``Lift_N``/``Drag_N`` row and an
        Internal Flow ``PressureDrop_Pa``/``ReynoldsNumber``/``FrictionFactor``
        row; there is no template-specific branching here.
        """
        c = self.cfg.columns

        def put(header: str, value, fmt: Optional[str] = None):
            if header not in self._col:      # defensive: never fabricate a
                return                       # column at write time
            cell = self._cell(exp.row, header)
            cell.value = value
            if fmt and value is not None:
                cell.number_format = fmt

        put(c.status, status)
        # Template-defined metrics.
        for metric_name, header in self._output_column_headers():
            value, fmt = self._metric_cell(res, metric_name)
            put(header, value, fmt)
        # Universal run bookkeeping (not a physical metric).
        put(c.iterations, res.iterations)
        put(c.converged, "YES" if res.converged else "NO")
        put(c.error, (res.error or "")[:500])
        put(c.started, res.started.strftime("%Y-%m-%d %H:%M:%S") if res.started else None)
        put(c.finished, res.finished.strftime("%Y-%m-%d %H:%M:%S") if res.finished else None)
        put(c.duration, res.duration_min, "0.00")
        put(c.case_dir, res.artifact_dir)
        # Legacy derived columns the workbook carries (e.g. FL/FD) — written
        # when the result can compute them.
        for attr, accessor, fmt, _depends in _DERIVED_COLUMNS:
            header = getattr(c, attr)
            if header not in self._col:
                continue
            value = getattr(res, accessor, None)
            if value is not None:
                put(header, round(float(value), 4), fmt)
        self.save()

    # ------------------------------------------------------------------ #
    # GUI-facing helpers (v0.8) — small, additive, reuse the column map
    # ------------------------------------------------------------------ #
    def wbp_names(self) -> List[str]:
        """Names of the extra Workbench-parameter columns (``WBP:`` prefix)."""
        return sorted(self._wbp_cols)

    def read_row_outputs(self, row: int) -> Dict[str, object]:
        """Raw output-cell values for one row (for the dataset table/charts).

        Keys are canonical (cl, cd, cl_cd, lift, drag, iterations, converged,
        error, duration, case_dir) — independent of the user's header names.
        This is the legacy External-Aerodynamics view; the template-driven
        read is :meth:`read_row_metrics`.
        """
        c = self.cfg.columns
        out: Dict[str, object] = {}
        for key, header in (("cl", c.cl), ("cd", c.cd), ("cl_cd", c.cl_cd),
                            ("lift", c.lift), ("drag", c.drag),
                            ("iterations", c.iterations),
                            ("converged", c.converged), ("error", c.error),
                            ("duration", c.duration), ("case_dir", c.case_dir)):
            col = self._col.get(header)
            out[key] = self.ws.cell(row=row, column=col).value if col else None
        return out

    def read_row_metrics(self, row: int) -> Dict[str, "MetricValue"]:
        """Generic output read for one row: the template's declared metric
        columns → :class:`MetricValue` dict, keyed by *template metric name* —
        the mirror image of :meth:`write_result`'s metric loop. Blank cells
        read as ``None`` values; unreadable cells are logged and treated as
        absent. The dict is ready for a generic :class:`CaseResult`.
        """
        out: Dict[str, MetricValue] = {}
        template = self._template()
        for metric_name, header in self._output_column_headers():
            col = self._col.get(header)
            raw = self.ws.cell(row=row, column=col).value if col else None
            md = template.metric(metric_name)
            out[metric_name] = MetricValue(
                metric_name, _as_float(raw), md.unit if md is not None else "")
        return out

    def update_input(self, row: int, field: str, value: float) -> None:
        """Edit one input cell of a *not-yet-run* row.

        ``field`` is ``"aoa"``, ``"velocity"``, or a WBP parameter name.
        The GUI is responsible for only offering this on PENDING/FAILED/SKIP
        rows; the manager just writes and saves.
        """
        c = self.cfg.columns
        if field == "aoa":
            col = self._col[c.aoa]
        elif field == "velocity":
            col = self._col[c.velocity]
        elif field in self._wbp_cols:
            col = self._wbp_cols[field]
        else:
            raise KeyError(f"Unknown input field '{field}'")
        self.ws.cell(row=row, column=col).value = float(value)
        self.save()

    def set_status(self, row: int, status: str) -> None:
        """Directly set the Status cell (e.g. toggle SKIP, clear to re-queue)."""
        self._cell(row, self.cfg.columns.status).value = status or None
        self.save()

    def append_experiment(self, aoa: float, velocity: float,
                          extra: Optional[Dict[str, float]] = None) -> int:
        """Add a new schedule row after the last populated one; returns its
        row number. Copies the input-cell number format from the row above so
        the sheet keeps looking hand-made."""
        c = self.cfg.columns
        last = self.cfg.header_row
        col_aoa, col_vel = self._col[c.aoa], self._col[c.velocity]
        for r in range(self.cfg.header_row + 1, self.ws.max_row + 1):
            if (self.ws.cell(row=r, column=col_aoa).value is not None
                    or self.ws.cell(row=r, column=col_vel).value is not None):
                last = r
        row = last + 1
        for col, value in ((col_aoa, float(aoa)), (col_vel, float(velocity))):
            cell = self.ws.cell(row=row, column=col, value=value)
            above = self.ws.cell(row=max(self.cfg.header_row + 1, row - 1),
                                 column=col)
            if above.number_format and above.number_format != "General":
                cell.number_format = above.number_format
        for name, value in (extra or {}).items():
            if name in self._wbp_cols:
                self.ws.cell(row=row, column=self._wbp_cols[name],
                             value=float(value))
        self.save()
        return row

    # ------------------------------------------------------------------ #
    # Crash-safe saving
    # ------------------------------------------------------------------ #
    def save(self) -> None:
        """Atomic save with retries.

        Two real-world failure modes are handled:
        * process killed mid-write  → write to a temp file, then ``os.replace``
          (atomic on the same filesystem), so the workbook is never truncated;
        * user has the file open in Excel (Windows write lock) → retry with a
          countdown message; after the retry budget, raise ExcelWriteError.
          Results are additionally persisted as JSON per case by the
          orchestrator, so nothing is lost either way.
        """
        tmp = self.path.with_suffix(".xlsx.tmp")
        last_exc: Optional[Exception] = None
        for attempt in range(1, self.cfg.save_retries + 1):
            try:
                self.wb.save(tmp)
                os.replace(tmp, self.path)
                return
            except PermissionError as exc:
                last_exc = exc
                log.warning(
                    "Workbook is locked (open in Excel?). Retry %d/%d in %.0fs — "
                    "please close '%s'.",
                    attempt, self.cfg.save_retries, self.cfg.save_retry_wait_s, self.path.name,
                )
                time.sleep(self.cfg.save_retry_wait_s)
            except Exception as exc:  # disk full, sync drive weirdness, ...
                last_exc = exc
                log.error("Workbook save failed (attempt %d): %s", attempt, exc)
                time.sleep(1.0)
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise ExcelWriteError(f"Could not save {self.path}: {last_exc}")

    # ------------------------------------------------------------------ #
    def dump_recovery_csv(self, path: Path, exp: Experiment, res: CaseResult, status: str) -> None:
        """Last-resort sidecar written when the workbook stays locked.

        Phase 8E: output column headers are template-driven (the template's
        declared metrics + universal bookkeeping) rather than hardcoded
        aero columns.
        """
        new = not path.exists()
        with open(path, "a", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            if new:
                # Build header row: row + template input params + status +
                # template metrics + bookkeeping.
                headers = ["row"]
                for name in self._study_io.input_parameter_names():
                    headers.append(name)
                headers.append("status")
                for _, h in self._output_column_headers():
                    headers.append(h)
                headers += ["iterations", "converged", "error"]
                w.writerow(headers)
            row_vals = [exp.row]
            for name in self._study_io.input_parameter_names():
                pv = exp.parameter(name)
                row_vals.append(pv.value if pv is not None else None)
            row_vals.append(status)
            for metric_name, _ in self._output_column_headers():
                mv = res.metric(metric_name)
                row_vals.append(mv.value if mv is not None else None)
            row_vals += [res.iterations, res.converged, res.error or ""]
            w.writerow(row_vals)
        log.error("Result for row %d written to recovery file %s — merge it into the "
                  "workbook manually or rerun after closing Excel.", exp.row, path)
