"""Stage 6 (T7) — realistic sample data for the stress-matrix screenshots.

Builds several *different* sample projects (never a single fixed dataset) so
the visual samples exercise the responsive workspace against real study
shapes:

    small-aero      default 8-row External Aerodynamics sweep, all DONE
    medium-aero     24-row External Aerodynamics sweep + one WBP column
                    (``WBP:FlapAngle`` — realistic metadata-driven extra
                    parameter), all DONE, realistic aero results
    large-queue     64-row External Aerodynamics sweep, all DONE — the Queue
                    genuinely overflows its height and must scroll
    mixed-status    12-row External Aerodynamics sweep with a realistic mix:
                    DONE / PENDING / FAILED(+Error) / SKIP
    internal-flow   8-row Internal Flow sweep (Inlet Velocity × Pipe
                    Diameter, from that template's metadata) — proves the
                    same GUI renders both templates with no template branch.
                    Results are pre-populated (the internal template has no
                    solver in this phase; the workbook is exactly what a
                    completed batch leaves behind — plain values).

Result rows are written as plain values, matching how the engine writes a
completed schedule (never formulas), so the loaded project is immediately
populated — no mock run needed.

Usage (imported by capture_stage6.py):
    import stage6_samples
    cfg_path = stage6_samples.build("medium-aero", tmp_root)
"""

from __future__ import annotations

import math
import random
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

# Standard result headers — the same set every completed schedule carries
# (config.ColumnMap). Sample rows fill these in.
_OUTPUT_HEADERS = ["Status", "CL", "CD", "CL/CD", "Lift_N", "Drag_N",
                   "Iterations", "Converged", "Error", "CaseDir",
                   "Duration_min"]

_RHO, _AREA = 1.225, 1.0          # External Aero reference values


def _aero(cl0: float, slope: float, stall: float,
          aoa: float, vel: float) -> Dict[str, float]:
    """Realistic thin-airfoil coefficients + lift/drag forces for one row."""
    cl = cl0 + slope * aoa
    if abs(aoa) > stall:                          # soft stall roll-off
        cl *= max(0.55, 1.0 - 0.06 * (abs(aoa) - stall))
    cd = 0.014 + 0.045 * cl * cl
    q = 0.5 * _RHO * vel * vel * _AREA
    return {"cl": cl, "cd": cd, "cl_cd": cl / cd,
            "lift": q * cl, "drag": q * cd}


def _internal(vel: float, dia: float, rho: float = 998.2,
              mu: float = 1.002e-3, length: float = 1.0) -> Dict[str, float]:
    """Realistic Internal Flow row: Darcy–Weisbach pressure drop, Reynolds
    number and friction factor from the pipe metadata. Filled into the
    standard result columns so the metadata-driven UI renders (output-column
    templating is a documented later generalization)."""
    re = rho * vel * dia / mu
    f = 64.0 / re if re < 2300 else 0.316 * re ** -0.25   # Blasius
    dP = f * (length / dia) * 0.5 * rho * vel * vel
    return {"cl": dP, "cd": f, "cl_cd": dP / f,
            "lift": re, "drag": dP * dia, "iterations": int(300 + 6 * re / 1000),
            "converged": "YES" if re < 1e6 else "NO"}


# --------------------------------------------------------------------------- #
# Workbook writers
# --------------------------------------------------------------------------- #
def _append_input_rows(ws, n_input_cols: int, rows: List[List[float]]) -> None:
    """Write additional input rows (matching the template's own input format)
    after the template's default rows."""
    first = 2
    for r, values in enumerate(rows, start=first):
        for c in range(1, n_input_cols + 1):
            ws.cell(row=r, column=c, value=float(values[c - 1]))


def _write_results(path: Path, rows: Dict[int, Dict[str, object]],
                   wbp_col: str | None = None) -> None:
    """Fill the standard result cells for the given sheet rows.

    ``rows`` maps Excel row number → {output-header: value} (headers without
    a value are left untouched). ``wbp_col`` optionally names a WBP column
    whose per-row value is passed in each record under ``"WBP:<name>"``.
    """
    wb = load_workbook(path)
    ws = wb["Experiments"]
    headers = {cell.value: cell.column
               for cell in ws[1] if cell.value is not None}
    for r, out in rows.items():
        for header, value in out.items():
            if value is None:
                continue
            col = headers.get(header)
            if col is None:
                continue
            ws.cell(row=r, column=col, value=value)
    wb.save(path)


# --------------------------------------------------------------------------- #
# Sample builders — each returns the config.yaml path
# --------------------------------------------------------------------------- #
def _config(tmp: Path, name: str, xlsx: Path, work: Path,
            template: str = "") -> Path:
    tpl = f"  template: \"{template}\"\n" if template else ""
    cfg = tmp / f"{name}.yaml"
    cfg.write_text(
        f"""
fluent:
  aoa_method: "geometry"
  wall_zones: ["wing"]
  reference: {{density: 1.225, area: 1.0}}
excel:
  file: "{xlsx.as_posix()}"
runtime:
  work_dir: "{work.as_posix()}"
  mock: true
{tpl}
""")
    return cfg


def _external_rows(aoas: List[float], vels: List[float]) -> List[List[float]]:
    """Cartesian grid, AOA varying slowest — matches the template's layout."""
    return [[a, v] for a in aoas for v in vels]


def build_small(tmp: Path) -> Path:
    from tools.make_experiment_template import build_template
    xlsx = tmp / "small.xlsx"
    build_template(xlsx)                       # default 8-row External Aero
    rows: Dict[int, Dict[str, object]] = {}
    rnd = random.Random(11)
    grid = _external_rows([0, 4, 8, 12], [20, 30])
    for i, (a, v) in enumerate(grid):
        r = 2 + i
        rho = _aero(0.35, 0.108, 12.0, a, v)
        rows[r] = {
            "Status": "DONE",
            "CL": round(rho["cl"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
            "CD": round(rho["cd"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
            "CL/CD": round(rho["cl_cd"], 4),
            "Lift_N": round(rho["lift"], 3), "Drag_N": round(rho["drag"], 3),
            "Iterations": int(420 + 30 * a + rnd.uniform(0, 60)),
            "Converged": "YES", "Duration_min": round(rnd.uniform(2.0, 6.0), 2),
        }
    _write_results(xlsx, rows)
    return _config(tmp, "small-aero", xlsx, tmp / "runs-small")


def build_medium(tmp: Path) -> Path:
    from tools.make_experiment_template import build_template
    xlsx = tmp / "medium.xlsx"
    build_template(xlsx)
    wb = load_workbook(xlsx)
    ws = wb["Experiments"]
    # Add a realistic Workbench extra-parameter column: WBP:FlapAngle.
    wbp_col = ws.max_column + 1
    hdr = ws.cell(row=1, column=wbp_col, value="WBP:FlapAngle")
    from openpyxl.styles import Font
    hdr.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    wb.save(xlsx)

    grid = _external_rows([0, 2, 4, 6, 8, 10], [15, 20, 25, 30])  # 24 rows
    _append_input_rows(ws, 2, grid)
    # add matching flap angles (same geometry groups share a mesh)
    for i in range(len(grid)):
        ws.cell(row=2 + i, column=wbp_col,
                value=float([0, 5, 10, 15][i % 4]))
    wb.save(xlsx)

    rows: Dict[int, Dict[str, object]] = {}
    rnd = random.Random(22)
    for i, (a, v) in enumerate(grid):
        r = 2 + i
        rho = _aero(0.35, 0.108, 12.0, a, v)
        rows[r] = {
            "Status": "DONE",
            "CL": round(rho["cl"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
            "CD": round(rho["cd"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
            "CL/CD": round(rho["cl_cd"], 4),
            "Lift_N": round(rho["lift"], 3), "Drag_N": round(rho["drag"], 3),
            "Iterations": int(420 + 30 * a + rnd.uniform(0, 60)),
            "Converged": "YES", "Duration_min": round(rnd.uniform(2.0, 8.0), 2),
        }
    _write_results(xlsx, rows)
    return _config(tmp, "medium-aero", xlsx, tmp / "runs-medium")


def build_large(tmp: Path) -> Path:
    from tools.make_experiment_template import build_template
    xlsx = tmp / "large.xlsx"
    build_template(xlsx)
    wb = load_workbook(xlsx)
    ws = wb["Experiments"]
    grid = _external_rows([0, 1, 2, 3, 4, 5, 6, 8], [10, 15, 20, 25, 30, 35,
                                                     40, 45])  # 64 rows
    _append_input_rows(ws, 2, grid)
    wb.save(xlsx)

    rows: Dict[int, Dict[str, object]] = {}
    rnd = random.Random(33)
    for i, (a, v) in enumerate(grid):
        r = 2 + i
        rho = _aero(0.35, 0.108, 12.0, a, v)
        rows[r] = {
            "Status": "DONE",
            "CL": round(rho["cl"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
            "CD": round(rho["cd"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
            "CL/CD": round(rho["cl_cd"], 4),
            "Lift_N": round(rho["lift"], 3), "Drag_N": round(rho["drag"], 3),
            "Iterations": int(420 + 30 * a + rnd.uniform(0, 60)),
            "Converged": "YES", "Duration_min": round(rnd.uniform(2.0, 8.0), 2),
        }
    _write_results(xlsx, rows)
    return _config(tmp, "large-queue", xlsx, tmp / "runs-large")


def build_mixed(tmp: Path) -> Path:
    from tools.make_experiment_template import build_template
    xlsx = tmp / "mixed.xlsx"
    build_template(xlsx)
    wb = load_workbook(xlsx)
    ws = wb["Experiments"]
    grid = _external_rows([0, 4, 8, 12, 16], [20, 30, 40])     # 15 rows
    _append_input_rows(ws, 2, grid)
    wb.save(xlsx)

    rows: Dict[int, Dict[str, object]] = {}
    rnd = random.Random(44)
    for i, (a, v) in enumerate(grid):
        r = 2 + i
        rho = _aero(0.35, 0.108, 12.0, a, v)
        status = "DONE"
        if i == 6:
            status = "FAILED"
        elif i == 9:
            status = "SKIP"
        elif i >= 12:
            status = "PENDING"                       # empty Status cell
        rec = {
            "Status": status,
            "Iterations": int(420 + 30 * a + rnd.uniform(0, 60)),
        }
        if status == "DONE":
            rec.update({
                "CL": round(rho["cl"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
                "CD": round(rho["cd"] * (1 + rnd.uniform(-0.004, 0.004)), 5),
                "CL/CD": round(rho["cl_cd"], 4),
                "Lift_N": round(rho["lift"], 3), "Drag_N": round(rho["drag"], 3),
                "Converged": "YES",
                "Duration_min": round(rnd.uniform(2.0, 8.0), 2),
            })
        elif status == "FAILED":
            rec["Error"] = ("Fluent diverged: continuity residual failed to "
                            "drop below 1e-4 after max iterations.")
        rows[r] = rec
    # PENDING rows keep an empty Status cell (the GUI maps empty → PENDING).
    _write_results(xlsx, rows)
    return _config(tmp, "mixed-status", xlsx, tmp / "runs-mixed")


def build_internal(tmp: Path) -> Path:
    from cfdauto.config import load_config
    from cfdauto.experiment_definition import ExperimentDefinition
    from tools.make_experiment_template import build_template
    xlsx = tmp / "internal.xlsx"
    cfg = _config(tmp, "internal-flow", xlsx, tmp / "runs-internal",
                  template="internal-flow")
    exp_def = ExperimentDefinition.for_config(load_config(cfg))
    build_template(xlsx, exp_def)               # Inlet Velocity × Pipe Diameter

    rows: Dict[int, Dict[str, object]] = {}
    for i, (v, d) in enumerate([(1.0, 0.05), (2.0, 0.05), (5.0, 0.05),
                                (10.0, 0.05), (1.0, 0.1), (2.0, 0.1),
                                (5.0, 0.1), (10.0, 0.1)]):
        r = 2 + i
        res = _internal(v, d)
        rows[r] = {
            "Status": "DONE",
            "CL": round(res["cl"], 2),
            "CD": round(res["cd"], 5),
            "CL/CD": round(res["cl_cd"], 2),
            "Lift_N": round(res["lift"], 0), "Drag_N": round(res["drag"], 2),
            "Iterations": res["iterations"], "Converged": res["converged"],
            "Duration_min": round(random.Random(i).uniform(3.0, 9.0), 2),
        }
    _write_results(xlsx, rows)
    return cfg


def build_demo(tmp: Path) -> Path:
    """Default 8-row External Aero schedule with NO results — the capture
    harness runs a mock batch on it so every panel (including Images, which
    needs real case-dir artefacts) is populated by the live path."""
    from tools.make_experiment_template import build_template
    xlsx = tmp / "demo.xlsx"
    build_template(xlsx)
    return _config(tmp, "demo", xlsx, tmp / "runs-demo")


# Registry of builders — keyed by sample name for the capture harness.
BUILDERS = {
    "small-aero": build_small,
    "medium-aero": build_medium,
    "large-queue": build_large,
    "mixed-status": build_mixed,
    "internal-flow": build_internal,
    "demo": build_demo,
}


def build(name: str, tmp: Path) -> Path:
    """Build one named sample project under ``tmp``; returns its config path."""
    builder = BUILDERS[name]
    if builder is None:
        raise KeyError(f"Unknown sample '{name}'; have {sorted(BUILDERS)}")
    return builder(tmp)
