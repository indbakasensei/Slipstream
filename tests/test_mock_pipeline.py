"""End-to-end pipeline test in mock mode + aero math checks.

Runs on any machine (no ANSYS): builds a schedule, executes it through the
real Orchestrator/ExcelManager/RunState with mock WB+Fluent backends, then
verifies resume semantics and the wind-axis vector math.

    pytest tests/ -q
"""

from __future__ import annotations

import math
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto import aero                                    # noqa: E402
from cfdauto.config import load_config                      # noqa: E402
from cfdauto.excel_manager import ExcelManager              # noqa: E402
from cfdauto.orchestrator import Orchestrator, build_controllers  # noqa: E402
from tools.make_experiment_template import build_template   # noqa: E402

CONFIG_TPL = """
fluent:
  baseline_case: ""
  aoa_method: "geometry"
  wall_zones: ["wing"]
  reference: {{density: 1.225, area: 1.0}}
solve:
  max_iterations: 100
excel:
  file: "{xlsx}"
runtime:
  work_dir: "{work}"
  mock: true
"""


@pytest.fixture()
def env(tmp_path: Path):
    xlsx = tmp_path / "experiments.xlsx"
    build_template(xlsx)
    cfg_file = tmp_path / "config.yaml"
    cfg_file.write_text(CONFIG_TPL.format(xlsx=xlsx.as_posix(),
                                          work=(tmp_path / "runs").as_posix()))
    return cfg_file, xlsx


def _run(cfg_file: Path, **kw) -> int:
    cfg = load_config(cfg_file)
    excel = ExcelManager(cfg.excel)
    wb, fl = build_controllers(cfg)
    return Orchestrator(cfg, excel, wb, fl).run(**kw)


def test_full_mock_pipeline_and_resume(env):
    cfg_file, xlsx = env

    failures = _run(cfg_file)
    assert failures == 0

    ws = load_workbook(xlsx)[
        "Experiments"]
    headers = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n_rows = ws.max_row - 1
    assert n_rows == 8                                   # 4 AOA x 2 velocities

    for r in range(2, ws.max_row + 1):
        assert ws.cell(r, headers["Status"]).value == "DONE"
        cl = ws.cell(r, headers["CL"]).value
        cd = ws.cell(r, headers["CD"]).value
        lift = ws.cell(r, headers["Lift_N"]).value
        v = ws.cell(r, headers["Velocity_m_s"]).value
        assert isinstance(cl, float) and isinstance(cd, float) and cd > 0
        assert ws.cell(r, headers["Converged"]).value == "YES"
        # Written CL/CD must equal the ratio of written columns.
        assert ws.cell(r, headers["CL/CD"]).value == pytest.approx(cl / cd, rel=1e-3)
        # Lift must be consistent with CL * q * A (the cross-check invariant).
        q = 0.5 * 1.225 * v * v
        assert lift == pytest.approx(cl * q * 1.0, rel=1e-2)

    # Mesh caching: 4 unique AOAs -> exactly 4 stored meshes for 8 rows.
    meshes = list((cfg_file.parent / "runs" / "meshes").iterdir())
    assert len(meshes) == 4

    # result.json exists for every case.
    cases = list((cfg_file.parent / "runs" / "cases").iterdir())
    assert len(cases) == 8
    assert all((c / "result.json").exists() for c in cases)

    # Resume: a second run must find nothing to do.
    assert _run(cfg_file) == 0
    cfg = load_config(cfg_file)
    assert ExcelManager(cfg.excel).pending(False, True) == []


def test_failed_rows_only_rerun_with_retry_flag(env):
    cfg_file, xlsx = env
    # Sabotage one row (negative velocity fails Experiment.validate()).
    from openpyxl import load_workbook as lw
    wb = lw(xlsx)
    ws = wb["Experiments"]
    ws.cell(3, 2).value = -5.0
    wb.save(xlsx)

    assert _run(cfg_file) == 1                            # exactly one failure
    ws = load_workbook(xlsx)["Experiments"]
    headers = {c.value: i + 1 for i, c in enumerate(ws[1])}
    assert ws.cell(3, headers["Status"]).value == "FAILED"
    assert "velocity" in str(ws.cell(3, headers["Error"]).value)

    assert _run(cfg_file) == 0                            # FAILED not re-queued
    # Fix it and retry.
    wb = lw(xlsx)
    wb["Experiments"].cell(3, 2).value = 25.0
    wb.save(xlsx)
    assert _run(cfg_file, retry_failed=True) == 0
    ws = load_workbook(xlsx)["Experiments"]
    assert ws.cell(3, headers["Status"]).value == "DONE"


# ------------------------------------------------------------------------- #
# Wind-axis math
# ------------------------------------------------------------------------- #
def test_wind_axes_geometry_mode_fixed():
    flow, drag, lift = aero.wind_axes(8.0, "geometry")
    assert flow == [1, 0, 0] and drag == [1, 0, 0] and lift == [0, 1, 0]


@pytest.mark.parametrize("a", [-10.0, 0.0, 5.0, 12.5])
def test_wind_axes_velocity_vector_rotation(a):
    flow, drag, lift = aero.wind_axes(a, "velocity_vector")
    r = math.radians(a)
    assert flow == pytest.approx([math.cos(r), math.sin(r), 0], abs=1e-9)
    assert drag == pytest.approx(flow, abs=1e-12)          # drag along flow
    assert lift == pytest.approx([-math.sin(r), math.cos(r), 0], abs=1e-9)
    # Orthogonality + unit length always hold.
    assert sum(d * l for d, l in zip(drag, lift)) == pytest.approx(0, abs=1e-9)
    assert sum(x * x for x in lift) == pytest.approx(1, abs=1e-9)


def test_coefficient_force_roundtrip():
    f = aero.force_from_coefficient(0.8, 1.225, 30.0, 1.5)
    assert aero.coefficient_from_force(f, 1.225, 30.0, 1.5) == pytest.approx(0.8)
