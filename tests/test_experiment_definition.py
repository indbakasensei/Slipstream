"""Universal CFD Platform, Phase 3B — unit tests for ExperimentDefinition
and template-driven workbook generation.

ExperimentDefinition is the runtime materialization of a StudyDefinition:
it produces concrete default rows, exposes the spreadsheet/editable/
validation views, and validates values against the template metadata.
These tests also lock in the byte-compatibility of the generated workbook
(same headers, same 8 example rows, same input formatting) — the single
most important guarantee of this sprint.
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto.experiment_definition import ExperimentDefinition   # noqa: E402
from cfdauto.platform import StudyDefinition                     # noqa: E402
from cfdauto.simulation_context import SimulationContext         # noqa: E402
from tools.make_experiment_template import build_template        # noqa: E402


# --------------------------------------------------------------------- #
# Group: runtime materialization from the template
# --------------------------------------------------------------------- #
def test_default_experiment_definition_exposes_the_input_schema():
    ed = ExperimentDefinition.default()
    assert ed.column_names() == ["AOA_deg", "Velocity_m_s"]
    assert ed.input_columns() == ["AOA", "Velocity"]
    assert ed.editable_columns() == ["AOA_deg", "Velocity_m_s"]
    assert ed.default_values() == {"AOA_deg": 0.0, "Velocity_m_s": 20.0}


def test_default_experiment_rows_are_the_cartesian_product_in_input_order():
    ed = ExperimentDefinition.default()
    rows = ed.default_experiment_rows()
    # AOA (0,4,8,12) × Velocity (20,30), AOA varying slowest — exactly the
    # 8 rows the legacy hardcoded template produced.
    assert rows == [
        (0.0, 20.0), (0.0, 30.0), (4.0, 20.0), (4.0, 30.0),
        (8.0, 20.0), (8.0, 30.0), (12.0, 20.0), (12.0, 30.0)]


def test_spreadsheet_columns_carry_full_ordered_metadata():
    cols = ExperimentDefinition.default().spreadsheet_columns()
    assert [c["column_name"] for c in cols] == ["AOA_deg", "Velocity_m_s"]
    assert [c["display_name"] for c in cols] == ["AOA", "Velocity"]
    assert [c["order"] for c in cols] == [0, 1]


def test_experiment_definition_references_not_copies_the_study():
    """Runtime state must not duplicate the platform metadata — it wraps it."""
    ctx = SimulationContext.default()
    ed = ExperimentDefinition.from_context(ctx)
    assert ed.study is ctx.study_definition


def test_empty_study_definition_yields_no_rows_without_crashing():
    ed = ExperimentDefinition(study=StudyDefinition())
    assert ed.default_experiment_rows() == []
    assert ed.column_names() == []
    assert ed.default_values() == {}


# --------------------------------------------------------------------- #
# Group: template-driven validation (delegates to ParameterDefinition)
# --------------------------------------------------------------------- #
def test_validate_value_uses_parameter_bounds():
    ed = ExperimentDefinition.default()
    assert ed.validate_value("AOA_deg", 5.0) == []
    assert ed.validate_value("AOA_deg", 200.0)          # above max 90
    assert ed.validate_value("Velocity_m_s", -1.0)       # below min 0.01
    # Unknown columns aren't this study's inputs → nothing to validate.
    assert ed.validate_value("WBP:Flap", 3.0) == []


def test_validate_row_reports_all_problems():
    ed = ExperimentDefinition.default()
    assert ed.validate_row({"AOA_deg": 5.0, "Velocity_m_s": 20.0}) == []
    problems = ed.validate_row({"AOA_deg": 200.0, "Velocity_m_s": 20.0})
    assert len(problems) == 1 and "AOA" in problems[0]


# --------------------------------------------------------------------- #
# Group: workbook compatibility (the sprint's key guarantee)
# --------------------------------------------------------------------- #
def test_generated_workbook_has_the_expected_headers_and_rows(tmp_path):
    path = tmp_path / "e.xlsx"
    build_template(path)
    ws = load_workbook(path)["Experiments"]
    headers = [ws.cell(1, c).value for c in range(1, 3)]
    assert headers == ["AOA_deg", "Velocity_m_s"]        # input headers unchanged
    assert ws.max_row == 9                                 # header + 8 example rows
    # The 8 example rows, in the exact legacy order.
    got = [(ws.cell(r, 1).value, ws.cell(r, 2).value) for r in range(2, 10)]
    assert got == [(0, 20), (0, 30), (4, 20), (4, 30),
                   (8, 20), (8, 30), (12, 20), (12, 30)]


def test_generated_input_cells_keep_their_formatting(tmp_path):
    path = tmp_path / "e.xlsx"
    build_template(path)
    ws = load_workbook(path)["Experiments"]
    aoa_cell = ws.cell(2, 1)
    assert aoa_cell.number_format == "0.0"
    assert aoa_cell.font.name == "Calibri"
    assert aoa_cell.fill.fgColor.rgb.endswith("DDEBF7")   # INPUT_FILL
    hdr = ws.cell(1, 1)
    assert hdr.font.bold is True
    assert hdr.fill.fgColor.rgb.endswith("1F4E79")        # HEADER_FILL


def test_generated_workbook_reads_back_through_excel_manager(tmp_path):
    """End-to-end: the template-generated workbook is consumed by the real
    ExcelManager exactly as before — 8 experiments, AOA/Velocity intact."""
    from cfdauto.config import ExcelConfig
    from cfdauto.excel_manager import ExcelManager
    path = tmp_path / "e.xlsx"
    build_template(path)
    exps = ExcelManager(ExcelConfig(file=str(path))).read_experiments()
    assert len(exps) == 8
    assert exps[0].aoa_deg == 0.0 and exps[0].velocity == 20.0
    assert exps[-1].aoa_deg == 12.0 and exps[-1].velocity == 30.0
