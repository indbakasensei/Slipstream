"""Universal CFD Platform, Phase 6 — Internal Flow template validation.

These tests prove the platform is domain-agnostic: a second, fundamentally
different CFD workflow (internal pipe flow) is defined purely as data and
driven through the *existing* runtime (ExperimentDefinition, StudyIO,
Experiment) with no core changes. They cover registration, study
generation, workbook generation, import round-trip, validation, template
isolation, and cross-template compatibility.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto.config import ColumnMap                        # noqa: E402
from cfdauto.experiment_definition import ExperimentDefinition  # noqa: E402
from cfdauto.platform import (                              # noqa: E402
    EXTERNAL_AERODYNAMICS,
    INTERNAL_FLOW,
    get_default_registry,
    get_default_template,
)
from cfdauto.simulation_context import SimulationContext     # noqa: E402
from cfdauto.study_io import StudyIO                         # noqa: E402
from tools.make_experiment_template import build_template    # noqa: E402


def _internal_flow_exp_def() -> ExperimentDefinition:
    """An ExperimentDefinition for Internal Flow, via the existing APIs."""
    ctx = SimulationContext(template=get_default_registry().get("internal-flow"))
    return ExperimentDefinition.from_context(ctx)


# --------------------------------------------------------------------- #
# Group: registration + coexistence (default unchanged)
# --------------------------------------------------------------------- #
def test_internal_flow_is_registered_alongside_external_aerodynamics():
    reg = get_default_registry()
    assert set(reg.ids()) == {"external-aerodynamics", "internal-flow"}
    assert reg.get("internal-flow") is INTERNAL_FLOW


def test_registering_internal_flow_did_not_change_the_default():
    # The critical coexistence guarantee: every existing runtime path still
    # resolves to External Aerodynamics.
    assert get_default_template() is EXTERNAL_AERODYNAMICS
    assert SimulationContext.default().template is EXTERNAL_AERODYNAMICS
    assert ExperimentDefinition.default().study is EXTERNAL_AERODYNAMICS.study_definition


# --------------------------------------------------------------------- #
# Group: template structure
# --------------------------------------------------------------------- #
def test_internal_flow_template_has_realistic_parameters_and_metrics():
    t = INTERNAL_FLOW
    assert {p.name for p in t.supported_parameters} == {
        "inlet_velocity", "fluid_density", "fluid_viscosity",
        "pipe_diameter", "pipe_length"}
    assert {m.name for m in t.supported_metrics} == {
        "pressure_drop", "reynolds_number", "friction_factor"}
    # Water-at-20C defaults, geometry driven via Workbench parameters.
    assert t.parameter("fluid_density").default_value == pytest.approx(998.2)
    assert t.parameter("pipe_diameter").workbench_parameter == "P1"
    assert t.parameter("inlet_velocity").workbench_parameter is None


# --------------------------------------------------------------------- #
# Group: study generation (no template-specific branching)
# --------------------------------------------------------------------- #
def test_experiment_definition_generates_internal_flow_study():
    ed = _internal_flow_exp_def()
    assert ed.column_names() == [
        "InletVelocity_m_s", "FluidDensity_kg_m3", "FluidViscosity_Pa_s",
        "PipeDiameter_m", "PipeLength_m"]
    rows = ed.default_experiment_rows()
    # inlet_velocity (4) x pipe_diameter (2), others fixed -> 8 rows.
    assert len(rows) == 8
    assert rows[0] == (1.0, 998.2, 0.001002, 0.05, 1.0)
    assert rows[-1] == (10.0, 998.2, 0.001002, 0.1, 1.0)


def test_build_experiment_produces_a_generic_internal_flow_experiment():
    ed = _internal_flow_exp_def()
    exp = ed.build_experiment(row=2, values={
        "inlet_velocity": 5.0, "fluid_density": 998.2,
        "fluid_viscosity": 0.001002, "pipe_diameter": 0.1, "pipe_length": 1.0})
    # Generic accessors work; the experiment has no airfoil parameters.
    assert exp.parameter("inlet_velocity").value == 5.0
    assert exp.parameter("pipe_diameter").value == 0.1
    assert "aoa" not in exp.parameters
    assert set(exp.parameters_dict()) == {
        "inlet_velocity", "fluid_density", "fluid_viscosity",
        "pipe_diameter", "pipe_length"}


# --------------------------------------------------------------------- #
# Group: workbook generation + import round-trip
# --------------------------------------------------------------------- #
def test_internal_flow_workbook_generation(tmp_path):
    ed = _internal_flow_exp_def()
    path = tmp_path / "internal_flow.xlsx"
    build_template(path, exp_def=ed)
    ws = load_workbook(path)["Experiments"]
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == ["InletVelocity_m_s", "FluidDensity_kg_m3",
                       "FluidViscosity_Pa_s", "PipeDiameter_m", "PipeLength_m"]
    assert ws.max_row == 9                            # header + 8 rows
    # First data row = first default row.
    assert [ws.cell(2, c).value for c in range(1, 6)] == [1, 998.2, 0.001002, 0.05, 1]


def test_internal_flow_workbook_round_trips_through_study_io(tmp_path):
    """Generate -> read back through StudyIO: the 8 rows reconstruct as
    generic Internal Flow experiments with the exact swept values."""
    ed = _internal_flow_exp_def()
    path = tmp_path / "internal_flow.xlsx"
    build_template(path, exp_def=ed)

    io = StudyIO(ed, ColumnMap())
    ws = load_workbook(path)["Experiments"]
    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    names = io.input_parameter_names()
    headers = io.input_column_headers()

    experiments = []
    for row in range(2, ws.max_row + 1):
        input_values = {name: ws.cell(row, header_to_col[hdr]).value
                        for name, hdr in zip(names, headers)}
        exp, warn = io.interpret_row(row, input_values, {}, "")
        assert warn is None
        experiments.append(exp)

    assert len(experiments) == 8
    swept = [(e.parameter("inlet_velocity").value, e.parameter("pipe_diameter").value)
             for e in experiments]
    assert swept == [(1, 0.05), (1, 0.1), (2, 0.05), (2, 0.1),
                     (5, 0.05), (5, 0.1), (10, 0.05), (10, 0.1)]


# --------------------------------------------------------------------- #
# Group: validation
# --------------------------------------------------------------------- #
def test_internal_flow_validation_uses_its_own_parameter_bounds():
    io = StudyIO(_internal_flow_exp_def(), ColumnMap())
    full = {"inlet_velocity": 2.0, "fluid_density": 998.2,
            "fluid_viscosity": 0.001002, "pipe_diameter": 0.05,
            "pipe_length": 1.0}
    assert io.validate_row(full) == []
    bad = dict(full, pipe_diameter=99.0)               # above max 10
    problems = io.validate_row(bad)
    assert len(problems) == 1 and "Pipe Diameter" in problems[0]


# --------------------------------------------------------------------- #
# Group: template isolation + cross-template compatibility
# --------------------------------------------------------------------- #
def test_templates_are_isolated_no_shared_parameters_or_metrics():
    ext_params = {p.name for p in EXTERNAL_AERODYNAMICS.supported_parameters}
    int_params = {p.name for p in INTERNAL_FLOW.supported_parameters}
    assert ext_params.isdisjoint(int_params)
    ext_metrics = {m.name for m in EXTERNAL_AERODYNAMICS.supported_metrics}
    int_metrics = {m.name for m in INTERNAL_FLOW.supported_metrics}
    assert ext_metrics.isdisjoint(int_metrics)


def test_both_templates_generate_independently_without_interference(tmp_path):
    ext = ExperimentDefinition.default()
    intf = _internal_flow_exp_def()
    # Two different studies, materialized side by side.
    assert ext.column_names() == ["AOA_deg", "Velocity_m_s"]
    assert intf.column_names()[0] == "InletVelocity_m_s"
    assert len(ext.default_experiment_rows()) == 8
    assert len(intf.default_experiment_rows()) == 8

    ep = tmp_path / "ext.xlsx"; build_template(ep)             # default (aero)
    ip = tmp_path / "int.xlsx"; build_template(ip, exp_def=intf)
    assert load_workbook(ep)["Experiments"].cell(1, 1).value == "AOA_deg"
    assert load_workbook(ip)["Experiments"].cell(1, 1).value == "InletVelocity_m_s"
