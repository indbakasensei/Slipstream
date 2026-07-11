"""Sprint 1 — behavioral tests for cfdauto.excel_manager.ExcelManager.

The Excel workbook *is* the run database (see the module's own docstring):
the Status column is the resume state machine, and save() is the only thing
standing between a finished case's results and total data loss if the user
has the file open. These tests target those two contracts plus the input
parsing that decides what "counts" as an experiment row.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto.config import ExcelConfig                   # noqa: E402
from cfdauto.exceptions import ConfigError, ExcelWriteError  # noqa: E402
from cfdauto.excel_manager import ExcelManager            # noqa: E402
from cfdauto.models import CaseResult, STATUS_DONE        # noqa: E402
from tools.make_experiment_template import build_template  # noqa: E402


def _headers(ws) -> dict:
    return {c.value: i + 1 for i, c in enumerate(ws[1])}


def _clear_rows(ws) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None


# --------------------------------------------------------------------- #
# Group: reading the schedule
#
# Regression Scenario: real schedules accumulate blank spacer rows,
# rows where the user filled in only one of AOA/velocity, and the
# occasional fat-fingered non-numeric cell. Any of these crashing the
# reader would take down the whole batch over one bad row.
# Expected Behaviour: only rows with both a valid AOA and a valid
# velocity become Experiments; everything else is skipped (with a
# warning), never raises.
# Why this test exists: read_experiments() is on the hot path of every
# single run — it must be forgiving of a hand-edited spreadsheet.
# --------------------------------------------------------------------- #
def test_read_experiments_skips_blank_half_filled_and_unparseable_rows(tmp_path):
    path = tmp_path / "e.xlsx"
    build_template(path)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    _clear_rows(ws)
    ws.cell(2, 1, 5.0); ws.cell(2, 2, 20.0)              # valid row
    # row 3 left entirely blank -> spacer row, silently skipped
    ws.cell(4, 1, 6.0)                                    # velocity missing
    ws.cell(5, 1, "not-a-number"); ws.cell(5, 2, 20.0)    # unparseable AOA
    wb.save(path)

    exps = ExcelManager(ExcelConfig(file=str(path))).read_experiments()
    assert len(exps) == 1
    assert exps[0].aoa_deg == 5.0 and exps[0].velocity == 20.0


# --------------------------------------------------------------------- #
# Group: resume semantics (pending())
#
# Regression Scenario: the Status column is the entire crash-recovery
# mechanism — a user kills the process mid-batch (RUNNING rows left
# behind), or reruns after fixing a FAILED row, or the sheet has a typo'd
# status string. Getting any of these wrong means either re-running
# already-DONE cases (wasted compute) or silently never re-running rows
# that need it.
# Expected Behaviour: PENDING/blank always included; RUNNING only when
# rerun_stale_running is set; FAILED only when retry_failed is set;
# DONE/SKIP never included; an unrecognized status is treated as pending
# (fails safe, not silent).
# Why this test exists: this is the resume contract the README's "kill
# the process anytime" guarantee depends on.
# --------------------------------------------------------------------- #
STATUSES = ["", "PENDING", "RUNNING", "FAILED", "DONE", "SKIP", "GARBAGE"]


def _cfg_with_statuses(tmp_path) -> ExcelConfig:
    path = tmp_path / "e.xlsx"
    build_template(path)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    _clear_rows(ws)
    h = _headers(ws)
    for i, status in enumerate(STATUSES):
        r = i + 2
        ws.cell(r, h["AOA_deg"], 5.0)
        ws.cell(r, h["Velocity_m_s"], 20.0)
        ws.cell(r, h["Status"], status)
    wb.save(path)
    return ExcelConfig(file=str(path))


@pytest.mark.parametrize("retry_failed,rerun_stale,expected", [
    (False, True, {"", "PENDING", "RUNNING", "GARBAGE"}),
    (False, False, {"", "PENDING", "GARBAGE"}),
    (True, True, {"", "PENDING", "RUNNING", "FAILED", "GARBAGE"}),
])
def test_pending_resume_semantics_by_status(tmp_path, retry_failed, rerun_stale, expected):
    cfg = _cfg_with_statuses(tmp_path)
    row_to_status = {i + 2: s for i, s in enumerate(STATUSES)}
    todo = ExcelManager(cfg).pending(retry_failed, rerun_stale)
    assert {row_to_status[e.row] for e in todo} == expected


# --------------------------------------------------------------------- #
# Group: writing results back (mark_running / write_result)
#
# Regression Scenario: the workbook is the sole visible record of a run
# for the user — a formatting bug (e.g. wrong number written, error text
# not truncated, timestamp mis-formatted) shows up directly in their
# spreadsheet, not in a log file they may never open.
# Expected Behaviour: after mark_running() + write_result(), reloading
# the file from disk shows exactly the values written, error text
# capped at 500 characters, and Converged rendered as YES/NO.
# Why this test exists: verifies the write path end-to-end through an
# actual save+reload, not just in-memory state.
# --------------------------------------------------------------------- #
def test_write_result_round_trips_through_disk_and_truncates_long_errors(tmp_path):
    from datetime import datetime, timedelta

    path = tmp_path / "e.xlsx"
    build_template(path)
    mgr = ExcelManager(ExcelConfig(file=str(path)))
    exp = mgr.read_experiments()[0]

    mgr.mark_running(exp, "runs/cases/r002_aoa0_v20")
    started = datetime(2026, 1, 1, 10, 0, 0)
    res = CaseResult(cl=0.812345, cd=0.0345, lift_n=120.5, drag_n=5.1,
                      iterations=450, converged=True,
                      started=started, finished=started + timedelta(minutes=5),
                      error="x" * 600, artifact_dir="runs/cases/r002_aoa0_v20")
    mgr.write_result(exp, res, STATUS_DONE)

    reloaded = load_workbook(path)["Experiments"]
    h = _headers(reloaded)
    assert reloaded.cell(exp.row, h["Status"]).value == "DONE"
    assert reloaded.cell(exp.row, h["CL"]).value == pytest.approx(0.812345)
    assert reloaded.cell(exp.row, h["Converged"]).value == "YES"
    assert reloaded.cell(exp.row, h["Started"]).value == "2026-01-01 10:00:00"
    assert len(reloaded.cell(exp.row, h["Error"]).value) == 500


# --------------------------------------------------------------------- #
# Group: extra Workbench-parameter (WBP:) columns
#
# Regression Scenario: a user adds a new design variable (e.g. flap
# angle) by adding a "WBP:FlapAngle" column — the documented, no-code
# extension mechanism. If this silently failed to round-trip, the extra
# parameter would never reach Workbench.
# Expected Behaviour: the column is recognized by its WBP: prefix and its
# value appears in Experiment.extra_wb_params.
# Why this test exists: protects the one supported way users extend the
# schedule without a code change.
# --------------------------------------------------------------------- #
def test_wbp_column_is_recognized_and_reaches_the_experiment(tmp_path):
    path = tmp_path / "e.xlsx"
    build_template(path)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    col = ws.max_column + 1
    ws.cell(1, col, "WBP:FlapAngle")
    ws.cell(2, col, 15.0)
    wb.save(path)

    mgr = ExcelManager(ExcelConfig(file=str(path)))
    assert mgr.wbp_names() == ["FlapAngle"]
    assert mgr.read_experiments()[0].extra_wb_params == {"FlapAngle": 15.0}


# --------------------------------------------------------------------- #
# Group: appending rows from the GUI ("add row")
#
# Regression Scenario: the GUI's "add experiment" action must append
# after the last real row, not at header+1 (which would silently
# overwrite an existing experiment), and should keep the sheet looking
# hand-made rather than reverting the new row to raw "General" format.
# Expected Behaviour: append_experiment() returns the correct new row
# number and copies the number format from the row above.
# Why this test exists: a hard-coded regression a user only notices when
# their new row overwrites row 2 by mistake.
# --------------------------------------------------------------------- #
def test_append_experiment_appends_after_last_row_and_keeps_number_format(tmp_path):
    path = tmp_path / "e.xlsx"
    build_template(path)                      # 8 pre-filled rows (rows 2-9)
    mgr = ExcelManager(ExcelConfig(file=str(path)))

    new_row = mgr.append_experiment(aoa=16.0, velocity=40.0)
    assert new_row == 10

    reloaded = load_workbook(path)["Experiments"]
    assert reloaded.cell(new_row, 1).value == 16.0
    assert reloaded.cell(new_row, 1).number_format == reloaded.cell(new_row - 1, 1).number_format


# --------------------------------------------------------------------- #
# Group: loud failure on schema mismatch
#
# Regression Scenario: config.yaml's column names (or sheet name) drift
# out of sync with the actual workbook — e.g. the user renamed a header
# or pointed the config at the wrong sheet. Proceeding anyway would mean
# writing results into the wrong columns or reading garbage.
# Expected Behaviour: ExcelManager construction raises ConfigError naming
# the missing column/sheet, before any row is read or written.
# Why this test exists: this is the schema contract between config.yaml
# and the workbook — it must fail at startup, not mid-batch.
# --------------------------------------------------------------------- #
def test_missing_required_column_or_sheet_raises_config_error(tmp_path):
    path = tmp_path / "e.xlsx"
    build_template(path)
    wb = load_workbook(path)
    wb["Experiments"].cell(1, 1).value = "Angle"   # rename AOA_deg away
    wb.save(path)
    with pytest.raises(ConfigError, match="AOA_deg"):
        ExcelManager(ExcelConfig(file=str(path)))

    path2 = tmp_path / "e2.xlsx"
    build_template(path2)
    with pytest.raises(ConfigError, match="Sheet"):
        ExcelManager(ExcelConfig(file=str(path2), sheet="DoesNotExist"))


# --------------------------------------------------------------------- #
# Group: crash-safe saving when the workbook is locked
#
# Regression Scenario: the single most common real-world failure this
# framework hits — the user has the workbook open in Excel (Windows file
# lock) when a case finishes and the engine tries to save results.
# Expected Behaviour: save() retries the configured number of times, then
# raises ExcelWriteError (never crashes the batch) and leaves no dangling
# .xlsx.tmp file or a truncated original workbook behind.
# Why this test exists: directly exercises the documented crash-safety
# guarantee in save()'s own docstring.
# --------------------------------------------------------------------- #
def test_save_retries_then_raises_and_cleans_up_when_locked(tmp_path, monkeypatch):
    path = tmp_path / "e.xlsx"
    build_template(path)
    mgr = ExcelManager(ExcelConfig(file=str(path), save_retries=3, save_retry_wait_s=0))

    attempts = []

    def fake_save(target):
        attempts.append(target)
        raise PermissionError("file is open in Excel")

    monkeypatch.setattr(mgr.wb, "save", fake_save)

    with pytest.raises(ExcelWriteError):
        mgr.save()
    assert len(attempts) == 3
    assert not path.with_suffix(".xlsx.tmp").exists()
    assert path.exists()


# --------------------------------------------------------------------- #
# Group: last-resort CSV sidecar
#
# Regression Scenario: the workbook stays locked for the entire retry
# budget (user never closes Excel) — this CSV is the only place the
# result survives.
# Expected Behaviour: the header is written exactly once; every call
# appends a data row.
# Why this test exists: a duplicated header (or a missing one on the
# first write) would make the recovery file useless to merge back in by
# hand, which is its entire purpose.
# --------------------------------------------------------------------- #
def test_dump_recovery_csv_writes_header_once_then_appends(tmp_path):
    path = tmp_path / "e.xlsx"
    build_template(path)
    mgr = ExcelManager(ExcelConfig(file=str(path)))
    exp = mgr.read_experiments()[0]
    res = CaseResult(cl=0.5, cd=0.05, converged=True)
    csv_path = tmp_path / "recovery.csv"

    mgr.dump_recovery_csv(csv_path, exp, res, STATUS_DONE)
    mgr.dump_recovery_csv(csv_path, exp, res, STATUS_DONE)

    lines = csv_path.read_text().splitlines()
    assert lines[0].startswith("row,aoa_deg")
    assert len(lines) == 3          # header + 2 data rows, no duplicate header
