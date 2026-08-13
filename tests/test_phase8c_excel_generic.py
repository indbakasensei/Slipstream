"""Phase 8C — Generic Excel / StudyIO boundary (v2.3.0-dev).

Phase 8B moved the generic *result* layer onto template-defined metrics
(``SimulationTemplate → output_columns() → CaseResult.metrics``). Phase 8C
moves that same contract into the workbook boundary:

    SimulationTemplate ──► input metadata ──► StudyIO / ExcelManager
        ──► workbook
    CaseResult.metrics ──► output metric metadata ──► StudyIO / ExcelManager
        ──► workbook

This suite proves the Excel layer is genuinely domain-neutral:

  * input columns resolve through the single StudyIO mechanism for every
    template (External Aero, Internal Flow, and a third canary) — there is
    no second input-column system;
  * output columns resolve from ``SimulationTemplate.output_columns()`` via
    StudyIO against the project's ColumnMap (no ``if template == …``);
  * legacy External Aero workbooks (CL / CD / CL/CD / Lift_N / Drag_N /
    FL/FD) are read and written exactly as before;
  * Internal Flow results write/read ``PressureDrop_Pa`` /
    ``ReynoldsNumber`` / ``FrictionFactor`` and never require aero fields;
  * round-trips (template → workbook → read → write → workbook) preserve
    metric names, values, units, case identity, and bookkeeping;
  * missing/unknown/blank/malformed/duplicate columns behave deterministically;
  * the Excel/StudyIO source contains no template-id branching.

Per the Phase 8C scope firewall the GUI, ledger, analytics, orchestrator,
and the NACA0012/CFD work are all untouched.
"""

from __future__ import annotations

import inspect
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto.config import ColumnMap, Config, ExcelConfig        # noqa: E402
from cfdauto.excel_manager import ExcelManager                    # noqa: E402
from cfdauto.execution.internal_flow import solve_internal_flow  # noqa: E402
from cfdauto.experiment_definition import ExperimentDefinition   # noqa: E402
from cfdauto.models import CaseResult, MetricValue, STATUS_DONE  # noqa: E402
from cfdauto.platform import EXTERNAL_AERODYNAMICS, INTERNAL_FLOW  # noqa: E402
from cfdauto.simulation_context import SimulationContext         # noqa: E402
from cfdauto.study_io import StudyIO                             # noqa: E402
from tools.make_experiment_template import build_template        # noqa: E402

# Reuse the Phase 8B test-only third-template canary (no production plugin,
# no solver) as the strongest proof that the Excel layer is domain-neutral.
from tests.test_phase8b_generic_metrics import CANARY_TEMPLATE   # noqa: E402


# --------------------------------------------------------------------------- #
# Fixtures — deterministic workbook fixtures following the project conventions
# (mirrors tools/make_experiment_template.py: input columns from the template's
# StudyDefinition, output columns from its declared metrics + universal
# bookkeeping, seeded with the template's default example sweep). These are
# NOT CFD results — fixture data for Excel/StudyIO validation only.
# --------------------------------------------------------------------------- #
def _generic_fixture_headers(template) -> List[str]:
    """Output headers for a Phase 8C fixture: Status + the template's metric
    columns + the universal bookkeeping columns (+ the legacy FL/FD derived
    column when the template declares the quantities it derives from)."""
    headers = ["Status"]
    headers += [h for _, h in template.output_columns()]
    if all(template.metric(m) is not None for m in ("lift", "drag")):
        headers.append("FL/FD")
    headers += ["Iterations", "Converged", "Error", "Started", "Finished",
                "Duration_min", "CaseDir"]
    return headers


def build_fixture_workbook(path: Path, template, *, result_row=None) -> Path:
    """Deterministic workbook fixture for any template (External Aero should
    use ``build_template`` — the canonical legacy fixture)."""
    ed = ExperimentDefinition.from_context(SimulationContext(template=template))
    input_headers = ed.column_names()
    headers = input_headers + _generic_fixture_headers(template)
    wb = Workbook()
    ws = wb.active
    ws.title = "Experiments"
    for col, name in enumerate(headers, start=1):
        ws.cell(1, col, name)
    for r, values in enumerate(ed.default_experiment_rows(), start=2):
        for i, v in enumerate(values):
            ws.cell(r, 1 + i, float(v))
    if result_row:
        for header, value in result_row.items():
            ws.cell(2, headers.index(header) + 1, value)
    doc = wb.create_sheet("ReadMe")
    doc.cell(1, 1, f"{template.id} fixture (Phase 8C)")
    wb.save(path)
    return path


def _internal_flow_fixture_row() -> Dict[str, object]:
    """Deterministic row-2 fixture values: water at V=1 m/s through D=0.05 m,
    L=1 m, from the analytical internal-flow solve. NOT CFD."""
    values = solve_internal_flow(
        {"inlet_velocity": 1.0, "fluid_density": 998.2,
         "fluid_viscosity": 1.002e-3, "pipe_diameter": 0.05,
         "pipe_length": 1.0})
    return {"Status": "DONE",
            "PressureDrop_Pa": round(values["pressure_drop"], 6),
            "ReynoldsNumber": round(values["reynolds_number"], 6),
            "FrictionFactor": round(values["friction_factor"], 6),
            "Iterations": 120, "Converged": "YES",
            "Started": "2026-08-01 09:00:00", "Finished": "2026-08-01 09:05:00",
            "Duration_min": 5.0, "CaseDir": "runs/cases/r002_inlet_velocity1"}


def _internal_flow_manager(path: Path) -> ExcelManager:
    cfg = Config()
    cfg.runtime.template = "internal-flow"
    cfg.excel.file = str(path)
    return ExcelManager.for_config(cfg)


def _canary_manager(path: Path) -> ExcelManager:
    cfg = Config()
    cfg.excel.file = str(path)
    io = StudyIO(ExperimentDefinition.from_context(
        SimulationContext(template=CANARY_TEMPLATE)), cfg.excel.columns)
    return ExcelManager(cfg.excel, study_io=io)


def _headers(ws) -> Dict[str, int]:
    return {c.value: i + 1 for i, c in enumerate(ws[1])}


# --------------------------------------------------------------------------- #
# 1–2 & 6–7. Input / output columns are template-driven (Phase 8C §5, §6)
# --------------------------------------------------------------------------- #
def test_external_aero_input_columns_are_template_driven():
    io = StudyIO.default(ColumnMap())
    assert io.input_parameter_names() == ["aoa", "velocity"]
    assert io.input_column_headers() == ["AOA_deg", "Velocity_m_s"]
    # A user ColumnMap rename flows through the same mechanism.
    assert StudyIO.default(ColumnMap(aoa="Alpha_deg", velocity="Vinf_m_s")
                           ).input_column_headers() == ["Alpha_deg", "Vinf_m_s"]


def test_internal_flow_input_columns_are_template_driven():
    io = StudyIO(ExperimentDefinition.from_context(
        SimulationContext(template=INTERNAL_FLOW)), ColumnMap())
    assert io.input_parameter_names() == [
        "inlet_velocity", "fluid_density", "fluid_viscosity",
        "pipe_diameter", "pipe_length"]
    assert io.input_column_headers() == [
        "InletVelocity_m_s", "FluidDensity_kg_m3", "FluidViscosity_Pa_s",
        "PipeDiameter_m", "PipeLength_m"]


def test_canary_input_columns_are_template_driven():
    io = StudyIO(ExperimentDefinition.from_context(
        SimulationContext(template=CANARY_TEMPLATE)), ColumnMap())
    assert io.input_column_headers() == ["TankRadius_m", "InletTemp_K"]


def test_external_aero_output_columns_match_legacy_layout(tmp_path):
    path = build_template(tmp_path / "e.xlsx")
    mgr = ExcelManager(ExcelConfig(file=str(path)))
    assert mgr._output_column_headers() == [
        ("cl", "CL"), ("cd", "CD"), ("l_over_d", "CL/CD"),
        ("lift", "Lift_N"), ("drag", "Drag_N")]
    # The full output schema reproduces the legacy ColumnMap layout exactly.
    assert mgr._output_headers() == ColumnMap().output_names()


def test_internal_flow_output_columns_declared(tmp_path):
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    mgr = _internal_flow_manager(path)
    assert mgr._output_column_headers() == [
        ("pressure_drop", "PressureDrop_Pa"),
        ("reynolds_number", "ReynoldsNumber"),
        ("friction_factor", "FrictionFactor")]
    assert mgr._output_headers() == [
        "Status", "PressureDrop_Pa", "ReynoldsNumber", "FrictionFactor",
        "Iterations", "Converged", "Error", "Started", "Finished",
        "Duration_min", "CaseDir"]
    # Internal Flow must not pull in the aero FL/FD derived column.
    assert "FL/FD" not in mgr._output_headers()


def test_canary_output_columns_have_no_aero_columns(tmp_path):
    path = build_fixture_workbook(tmp_path / "c.xlsx", CANARY_TEMPLATE)
    mgr = _canary_manager(path)
    assert mgr._output_column_headers() == [
        ("heat_rate", "HeatRate_W"),         # explicit output_column
        ("efficiency", "Efficiency"),        # display-name fallback
        ("vapor_fraction", "Vapor Fraction")]
    assert "CL" not in mgr._output_headers()


def test_output_metric_columns_honour_column_map_renames():
    # The metric→ColumnMap resolution honours user renames — including the
    # l_over_d ↔ cl_cd name translation (metric name ≠ ColumnMap attr).
    io = StudyIO.default(ColumnMap(cl="Clift", cl_cd="Eff"))
    assert dict(io.output_metric_columns()) == {
        "cl": "Clift", "cd": "CD", "l_over_d": "Eff",
        "lift": "Lift_N", "drag": "Drag_N"}


# --------------------------------------------------------------------------- #
# 3–4. External Aero legacy workbook read + result write (Phase 8C §4, §7)
# --------------------------------------------------------------------------- #
def test_external_aero_legacy_workbook_read(tmp_path):
    path = build_template(tmp_path / "e.xlsx")
    mgr = ExcelManager(ExcelConfig(file=str(path)))
    exps = mgr.read_experiments()
    assert [(e.aoa_deg, e.velocity) for e in exps] == [
        (0, 20), (0, 30), (4, 20), (4, 30), (8, 20), (8, 30), (12, 20), (12, 30)]
    # The legacy read_row_outputs dict keys are unchanged (GUI/analytics).
    assert set(mgr.read_row_outputs(2)) == {
        "cl", "cd", "cl_cd", "lift", "drag", "iterations", "converged",
        "error", "duration", "case_dir"}


def test_external_aero_result_write(tmp_path):
    path = build_template(tmp_path / "e.xlsx")
    mgr = ExcelManager(ExcelConfig(file=str(path)))
    exp = mgr.read_experiments()[0]
    started = datetime(2026, 1, 1, 10, 0, 0)
    res = CaseResult(template=EXTERNAL_AERODYNAMICS, cl=0.812345, cd=0.0345,
                     lift_n=120.5, drag_n=5.1, iterations=450, converged=True,
                     started=started, finished=started + timedelta(minutes=5),
                     error="x" * 600, artifact_dir="runs/cases/r002_aoa0_v20")
    mgr.write_result(exp, res, STATUS_DONE)
    reloaded = load_workbook(path)["Experiments"]
    h = _headers(reloaded)
    assert reloaded.cell(2, h["Status"]).value == "DONE"
    assert reloaded.cell(2, h["CL"]).value == pytest.approx(0.812345)
    assert reloaded.cell(2, h["CD"]).value == pytest.approx(0.0345)
    assert reloaded.cell(2, h["CL/CD"]).value == pytest.approx(
        round(0.812345 / 0.0345, 4))
    assert reloaded.cell(2, h["Lift_N"]).value == pytest.approx(120.5)
    assert reloaded.cell(2, h["Drag_N"]).value == pytest.approx(5.1)
    assert reloaded.cell(2, h["FL/FD"]).value == pytest.approx(
        round(120.5 / 5.1, 4))
    assert reloaded.cell(2, h["Converged"]).value == "YES"
    assert reloaded.cell(2, h["Started"]).value == "2026-01-01 10:00:00"
    assert len(reloaded.cell(2, h["Error"]).value) == 500


# --------------------------------------------------------------------------- #
# 5. External Aero semantic round-trip (Phase 8C §13–§14)
# --------------------------------------------------------------------------- #
def test_external_aero_semantic_round_trip(tmp_path):
    # A legacy-format workbook with a completed row → read → generic
    # representation → write → reload: the semantic contract survives.
    path = build_template(tmp_path / "e.xlsx")
    wb = load_workbook(path)
    ws = wb["Experiments"]
    h = _headers(ws)
    known = {"CL": 0.5, "CD": 0.05, "CL/CD": 10.0, "Lift_N": 50.0, "Drag_N": 5.0,
             "FL/FD": 10.0, "Iterations": 300, "Converged": "YES",
             "Started": "2026-01-01 09:00:00", "Finished": "2026-01-01 09:30:00",
             "Duration_min": 30.0, "CaseDir": "runs/cases/r002_aoa0_v20"}
    for header, value in known.items():
        ws.cell(2, h[header], value)
    wb.save(path)

    mgr = ExcelManager(ExcelConfig(file=str(path)))
    exp = mgr.read_experiments()[0]
    metrics = mgr.read_row_metrics(2)
    assert {k: v.value for k, v in metrics.items()} == {
        "cl": 0.5, "cd": 0.05, "l_over_d": 10.0, "lift": 50.0, "drag": 5.0}

    res = CaseResult(template=EXTERNAL_AERODYNAMICS, case_id=exp.case_id,
                     parameters=exp.parameters_dict(), metrics=metrics,
                     iterations=300, converged=True,
                     started=datetime(2026, 1, 1, 9, 0, 0),
                     finished=datetime(2026, 1, 1, 9, 30, 0),
                     artifact_dir="runs/cases/r002_aoa0_v20")
    mgr.write_result(exp, res, STATUS_DONE)

    reloaded = ExcelManager(ExcelConfig(file=str(path)))
    back = reloaded.read_row_metrics(2)
    assert back["cl"].value == pytest.approx(0.5)
    assert back["cd"].value == pytest.approx(0.05)
    assert back["l_over_d"].value == pytest.approx(10.0)
    assert back["lift"].value == pytest.approx(50.0)
    assert back["drag"].value == pytest.approx(5.0)
    assert back["lift"].unit == "N" and back["drag"].unit == "N"


# --------------------------------------------------------------------------- #
# 8–10. Internal Flow workbook read / write / semantic round-trip (§10, §15)
# --------------------------------------------------------------------------- #
def test_internal_flow_workbook_read(tmp_path):
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW,
                                  result_row=_internal_flow_fixture_row())
    mgr = _internal_flow_manager(path)
    exps = mgr.read_experiments()
    assert len(exps) == 8
    assert exps[0].parameters["inlet_velocity"].value == 1.0
    assert exps[0].parameters["pipe_diameter"].value == 0.05
    # No airfoil slot exists on an Internal Flow experiment.
    assert "aoa" not in exps[0].parameters and "velocity" not in exps[0].parameters


def test_internal_flow_result_write_and_read(tmp_path):
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    mgr = _internal_flow_manager(path)
    exp = mgr.read_experiments()[0]
    res = CaseResult(
        template=INTERNAL_FLOW, case_id=exp.case_id,
        parameters=exp.parameters_dict(),
        metrics={"pressure_drop": MetricValue("pressure_drop", 711.16, "Pa"),
                 "reynolds_number": MetricValue("reynolds_number", 99620.8, ""),
                 "friction_factor": MetricValue("friction_factor", 0.017809, "")},
        iterations=120, converged=True, artifact_dir="runs/cases/r002")
    mgr.write_result(exp, res, STATUS_DONE)

    m = mgr.read_row_metrics(2)
    assert {k: v.value for k, v in m.items()} == {
        "pressure_drop": 711.16, "reynolds_number": 99620.8,
        "friction_factor": 0.017809}
    assert m["pressure_drop"].unit == "Pa"
    assert "cl" not in m and "cd" not in m and "lift" not in m and "drag" not in m


def test_internal_flow_semantic_round_trip(tmp_path):
    # Deterministic fixture row → read → write back → read again: metrics,
    # units, case identity, and bookkeeping survive; no aero is required.
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW,
                                  result_row=_internal_flow_fixture_row())
    mgr = _internal_flow_manager(path)
    exp = mgr.read_experiments()[0]
    assert exp.case_id.startswith("r002_")

    metrics = mgr.read_row_metrics(2)
    assert set(metrics) == {"pressure_drop", "reynolds_number", "friction_factor"}
    out = mgr.read_row_outputs(2)
    assert out["iterations"] == 120 and out["converged"] == "YES"

    res = CaseResult(template=INTERNAL_FLOW, case_id=exp.case_id,
                     parameters=exp.parameters_dict(), metrics=metrics,
                     iterations=out["iterations"], converged=True,
                     artifact_dir="runs/cases/r002_inlet_velocity1")
    mgr.write_result(exp, res, STATUS_DONE)

    back = _internal_flow_manager(path)
    m2 = back.read_row_metrics(2)
    assert m2["pressure_drop"].value == pytest.approx(
        metrics["pressure_drop"].value, rel=1e-9)
    assert m2["reynolds_number"].value == pytest.approx(
        metrics["reynolds_number"].value, rel=1e-9)
    assert m2["friction_factor"].value == pytest.approx(
        metrics["friction_factor"].value, rel=1e-9)
    assert m2["pressure_drop"].unit == "Pa"
    assert back.read_row_outputs(2)["converged"] == "YES"


def test_internal_flow_write_does_not_touch_aero_columns(tmp_path):
    # Writing an Internal Flow result must never fabricate aero columns.
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    mgr = _internal_flow_manager(path)
    exp = mgr.read_experiments()[0]
    res = CaseResult(template=INTERNAL_FLOW,
                     metrics={"pressure_drop": MetricValue("pressure_drop", 1.0, "Pa")},
                     converged=True)
    mgr.write_result(exp, res, STATUS_DONE)
    reloaded = load_workbook(path)["Experiments"]
    for header in ("CL", "CD", "Lift_N", "Drag_N", "FL/FD"):
        assert header not in _headers(reloaded)


# --------------------------------------------------------------------------- #
# 11–12. Missing required inputs / unknown extra columns (Phase 8C §16)
# --------------------------------------------------------------------------- #
def test_missing_required_input_column_raises_config_error(tmp_path):
    # External Aero
    path = build_template(tmp_path / "e.xlsx")
    wb = load_workbook(path)
    wb["Experiments"].cell(1, 1).value = "Angle"
    wb.save(path)
    with pytest.raises(Exception, match="AOA_deg"):
        ExcelManager(ExcelConfig(file=str(path)))

    # Internal Flow — the same loud schema check for a non-aero template.
    path2 = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    wb2 = load_workbook(path2)
    wb2["Experiments"].cell(1, 1).value = "Vel"
    wb2.save(path2)
    with pytest.raises(Exception, match="InletVelocity_m_s"):
        _internal_flow_manager(path2)


def test_unknown_extra_column_is_preserved(tmp_path):
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    col = ws.max_column + 1
    ws.cell(1, col, "Notes")
    ws.cell(2, col, "hand-written note")
    wb.save(path)

    mgr = _internal_flow_manager(path)
    exp = mgr.read_experiments()[0]
    res = CaseResult(template=INTERNAL_FLOW,
                     metrics={"pressure_drop": MetricValue("pressure_drop", 1.0, "Pa")},
                     converged=True)
    mgr.write_result(exp, res, STATUS_DONE)

    reloaded = load_workbook(path)["Experiments"]
    assert reloaded.cell(2, col).value == "hand-written note"


# --------------------------------------------------------------------------- #
# 13. Deterministic column ordering (Phase 8C §17)
# --------------------------------------------------------------------------- #
def test_column_ordering_is_deterministic_and_matches_legacy(tmp_path):
    # External Aero: the template-driven schema reproduces ColumnMap order.
    ea = build_template(tmp_path / "e.xlsx")
    assert ExcelManager(ExcelConfig(file=str(ea)))._output_headers() == \
        ColumnMap().output_names()

    # Internal Flow: no alphabetical sort — template order is authoritative.
    iff = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    mgr = _internal_flow_manager(iff)
    first = mgr._output_headers()
    second = _internal_flow_manager(iff)._output_headers()
    assert first == second                       # deterministic across builds
    assert first == ["Status", "PressureDrop_Pa", "ReynoldsNumber",
                     "FrictionFactor", "Iterations", "Converged", "Error",
                     "Started", "Finished", "Duration_min", "CaseDir"]


# --------------------------------------------------------------------------- #
# 14. Workbench parameters survive the write path (Phase 8C §18)
# --------------------------------------------------------------------------- #
def test_workbench_parameter_preserved_across_write(tmp_path):
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    col = ws.max_column + 1
    ws.cell(1, col, "WBP:FlapAngle")
    ws.cell(2, col, 15.0)
    wb.save(path)

    mgr = _internal_flow_manager(path)
    exp = mgr.read_experiments()[0]
    # WBP params ride inside the parameter set marked source="wbp"
    # (the generic "extra parameters" marker — no aero field implied).
    assert exp.parameters["FlapAngle"].value == 15.0
    assert exp.parameters["FlapAngle"].source == "wbp"
    res = CaseResult(template=INTERNAL_FLOW,
                     metrics={"pressure_drop": MetricValue("pressure_drop", 1.0, "Pa")},
                     converged=True)
    mgr.write_result(exp, res, STATUS_DONE)

    reloaded = _internal_flow_manager(path)
    again = reloaded.read_experiments()[0]
    assert again.parameters["FlapAngle"].value == 15.0
    assert again.parameters["FlapAngle"].source == "wbp"


# --------------------------------------------------------------------------- #
# 15. Third-template canary through the real Excel layer (Phase 8C §12)
# --------------------------------------------------------------------------- #
def test_canary_workbook_round_trip(tmp_path):
    path = build_fixture_workbook(tmp_path / "c.xlsx", CANARY_TEMPLATE)
    mgr = _canary_manager(path)
    exps = mgr.read_experiments()
    assert len(exps) == 4
    assert exps[0].parameters["tank_radius"].value == 0.5

    res = CaseResult(
        template=CANARY_TEMPLATE, case_id="c001_tank_radius0.5",
        parameters=exps[0].parameters_dict(),
        metrics={"heat_rate": MetricValue("heat_rate", 45.7, "W"),
                 "efficiency": MetricValue("efficiency", 0.83, ""),
                 "vapor_fraction": MetricValue("vapor_fraction", 0.02, "")},
        converged=True, artifact_dir="runs/cases/c001")
    mgr.write_result(exps[0], res, STATUS_DONE)

    m = mgr.read_row_metrics(2)
    assert {k: v.value for k, v in m.items()} == {
        "heat_rate": 45.7, "efficiency": 0.83, "vapor_fraction": 0.02}
    assert m["heat_rate"].unit == "W"
    assert "cl" not in m and "lift" not in m and "drag" not in m


# --------------------------------------------------------------------------- #
# 16. Missing optional / blank / malformed / duplicate columns (§16)
# --------------------------------------------------------------------------- #
def test_missing_optional_output_columns_are_auto_created(tmp_path):
    # A minimal workbook with only the required inputs + Status: ExcelManager
    # creates the template's metric + bookkeeping columns at the right edge.
    ed = ExperimentDefinition.from_context(SimulationContext(template=INTERNAL_FLOW))
    path = tmp_path / "if_min.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Experiments"
    for col, name in enumerate(ed.column_names() + ["Status"], start=1):
        ws.cell(1, col, name)
    for r, values in enumerate(ed.default_experiment_rows(), start=2):
        for i, v in enumerate(values):
            ws.cell(r, 1 + i, float(v))
    wb.save(path)

    mgr = _internal_flow_manager(path)
    for header in mgr._output_headers():
        assert header in mgr._col
    # A write end-to-end works on the auto-created columns.
    exp = mgr.read_experiments()[0]
    res = CaseResult(template=INTERNAL_FLOW,
                     metrics={"pressure_drop": MetricValue("pressure_drop", 1.0, "Pa")},
                     converged=True)
    mgr.write_result(exp, res, STATUS_DONE)
    assert mgr.read_row_metrics(2)["pressure_drop"].value == 1.0


def test_blank_output_cell_reads_as_none(tmp_path):
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    mgr = _internal_flow_manager(path)
    metrics = mgr.read_row_metrics(2)
    assert all(v.value is None for v in metrics.values())
    assert metrics["pressure_drop"].unit == "Pa"   # unit from the template


def test_malformed_output_cell_reads_as_none(tmp_path):
    path = build_fixture_workbook(tmp_path / "if.xlsx", INTERNAL_FLOW)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    h = _headers(ws)
    ws.cell(2, h["PressureDrop_Pa"], "not-a-number")
    wb.save(path)
    mgr = _internal_flow_manager(path)
    # Unreadable output cells are treated as absent — never a crash.
    assert mgr.read_row_metrics(2)["pressure_drop"].value is None


def test_duplicate_header_is_deterministic(tmp_path):
    # A workbook with a duplicate output header: the last occurrence wins
    # (the existing _col mapping rule) for both read and write.
    path = build_template(tmp_path / "e.xlsx")
    wb = load_workbook(path)
    ws = wb["Experiments"]
    h = _headers(ws)
    dup = ws.max_column + 1
    ws.cell(1, dup, "CL")
    ws.cell(2, dup, 0.99)
    wb.save(path)

    mgr = ExcelManager(ExcelConfig(file=str(path)))
    assert mgr.read_row_metrics(2)["cl"].value == pytest.approx(0.99)
    # Write also targets the last occurrence.
    mgr.write_result(mgr.read_experiments()[0],
                     CaseResult(template=EXTERNAL_AERODYNAMICS, cl=0.5, cd=0.05,
                                converged=True), STATUS_DONE)
    reloaded = load_workbook(path)["Experiments"]
    assert reloaded.cell(2, dup).value == pytest.approx(0.5)


# --------------------------------------------------------------------------- #
# Static — no template-id branching in the generic Excel/StudyIO layer
# --------------------------------------------------------------------------- #
def test_generic_excel_layer_has_no_template_branching():
    import cfdauto.excel_manager
    import cfdauto.study_io
    forbidden = (
        'if template == "external-aerodynamics"',
        'if template == "internal-flow"',
        'if template.id == "external-aerodynamics"',
        'if template.id == "internal-flow"',
        'template == "external-aerodynamics"',
        'template == "internal-flow"',
    )
    for module in (cfdauto.excel_manager, cfdauto.study_io):
        src = inspect.getsource(module)
        for pattern in forbidden:
            assert pattern not in src, \
                f"{module.__name__} must not contain: {pattern}"
