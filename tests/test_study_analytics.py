"""Sprint 3 — behavioral tests for cfdauto.study_analytics.

Study Analytics is the first read-only summary computed after a batch
finishes. These tests protect three things: that the aggregate counts and
"best case" metrics are computed correctly (including the messy real-world
cases — missing values, rows that never finished, unconverged successes),
that tie-breaking is deterministic rather than incidental to dict/iteration
order, and that the module stays purely computational (no logging) while
its wiring into the orchestrator behaves exactly as documented (summary
lifecycle resets between runs, retries never leak between studies).
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto.config import ExcelConfig, load_config              # noqa: E402
from cfdauto.events import EventBus                               # noqa: E402
from cfdauto.excel_manager import ExcelManager                    # noqa: E402
from cfdauto.exceptions import DivergedError                      # noqa: E402
from cfdauto.models import CaseResult, STATUS_DONE                # noqa: E402
from cfdauto.orchestrator import Orchestrator                     # noqa: E402
from cfdauto.study_analytics import (                              # noqa: E402
    StudySummary,
    WarningCode,
    analyze_study,
)
from tools.make_experiment_template import build_template          # noqa: E402


def _headers(ws) -> dict:
    return {c.value: i + 1 for i, c in enumerate(ws[1])}


def _clear_rows(ws) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def _build(tmp_path, rows: list[dict]) -> ExcelManager:
    """Build a workbook with one row per dict in ``rows``. Each dict may set
    any of: aoa, velocity, status, cl_cd, lift, drag, iterations, converged."""
    path = tmp_path / "e.xlsx"
    build_template(path)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    _clear_rows(ws)
    h = _headers(ws)
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, h["AOA_deg"], r.get("aoa", 5.0))
        ws.cell(row, h["Velocity_m_s"], r.get("velocity", 20.0))
        if "status" in r:
            ws.cell(row, h["Status"], r["status"])
        if "cl_cd" in r:
            ws.cell(row, h["CL/CD"], r["cl_cd"])
        if "lift" in r:
            ws.cell(row, h["Lift_N"], r["lift"])
        if "drag" in r:
            ws.cell(row, h["Drag_N"], r["drag"])
        if "iterations" in r:
            ws.cell(row, h["Iterations"], r["iterations"])
        if "converged" in r:
            ws.cell(row, h["Converged"], r["converged"])
    wb.save(path)
    return ExcelManager(ExcelConfig(file=str(path)))


# --------------------------------------------------------------------- #
# Group: empty study
# --------------------------------------------------------------------- #
def test_empty_study_returns_zero_totals_and_a_single_warning(tmp_path):
    excel = _build(tmp_path, [])
    summary = analyze_study(excel, rows=[])
    assert summary.total_cases == 0
    assert summary.successful_cases == 0
    assert summary.failed_cases == 0
    assert len(summary.warnings) == 1
    assert summary.warnings[0].code == WarningCode.EMPTY_STUDY


# --------------------------------------------------------------------- #
# Group: all-success study
# --------------------------------------------------------------------- #
def test_all_success_computes_best_metrics_with_no_warnings(tmp_path):
    excel = _build(tmp_path, [
        {"status": STATUS_DONE, "cl_cd": 10.0, "lift": 50.0, "drag": 5.0,
         "iterations": 400, "converged": "YES"},
        {"status": STATUS_DONE, "cl_cd": 15.0, "lift": 80.0, "drag": 4.0,
         "iterations": 300, "converged": "YES"},
        {"status": STATUS_DONE, "cl_cd": 12.0, "lift": 60.0, "drag": 6.0,
         "iterations": 500, "converged": "YES"},
    ])
    summary = analyze_study(excel, rows=[2, 3, 4])
    assert summary.total_cases == 3
    assert summary.successful_cases == 3
    assert summary.failed_cases == 0
    assert summary.best_l_over_d == 15.0 and summary.best_l_over_d_row == 3
    assert summary.highest_lift_n == 80.0 and summary.highest_lift_row == 3
    assert summary.lowest_drag_n == 4.0 and summary.lowest_drag_row == 3
    assert (summary.fastest_convergence_iterations == 300
           and summary.fastest_convergence_row == 3)
    assert summary.warnings == []


# --------------------------------------------------------------------- #
# Group: mixed success/failure
# --------------------------------------------------------------------- #
def test_mixed_success_and_failure_counts_and_warns(tmp_path):
    excel = _build(tmp_path, [
        {"status": STATUS_DONE, "cl_cd": 10.0, "lift": 50.0, "drag": 5.0,
         "iterations": 400, "converged": "YES"},
        {"status": "FAILED"},
        {"status": "FAILED"},
    ])
    summary = analyze_study(excel, rows=[2, 3, 4])
    assert summary.successful_cases == 1
    assert summary.failed_cases == 2
    codes = [w.code for w in summary.warnings]
    assert WarningCode.CASE_FAILED in codes
    failed_warning = next(w for w in summary.warnings if w.code == WarningCode.CASE_FAILED)
    assert "2" in failed_warning.message


# --------------------------------------------------------------------- #
# Group: rows that never finished (stable, per-row warning rules)
# --------------------------------------------------------------------- #
def test_pending_and_running_rows_produce_per_row_warnings(tmp_path):
    excel = _build(tmp_path, [
        {"status": STATUS_DONE, "cl_cd": 10.0, "lift": 50.0, "drag": 5.0,
         "iterations": 400, "converged": "YES"},
        {"status": "RUNNING"},
        {"status": ""},          # blank == pending
    ])
    summary = analyze_study(excel, rows=[2, 3, 4])
    assert summary.successful_cases == 1
    assert summary.failed_cases == 0
    by_code = {w.code: w for w in summary.warnings}
    assert WarningCode.ROW_STILL_RUNNING in by_code
    assert "Row 3" in by_code[WarningCode.ROW_STILL_RUNNING].message
    assert WarningCode.ROW_STILL_PENDING in by_code
    assert "Row 4" in by_code[WarningCode.ROW_STILL_PENDING].message


# --------------------------------------------------------------------- #
# Group: missing / non-numeric values never crash the summary
# --------------------------------------------------------------------- #
def test_missing_numeric_values_are_skipped_without_crashing(tmp_path):
    excel = _build(tmp_path, [
        {"status": STATUS_DONE},                          # every metric blank
        {"status": STATUS_DONE, "cl_cd": "N/A", "lift": None, "drag": "x"},
    ])
    summary = analyze_study(excel, rows=[2, 3])
    assert summary.successful_cases == 2
    assert summary.best_l_over_d is None
    assert summary.highest_lift_n is None
    assert summary.lowest_drag_n is None
    assert summary.fastest_convergence_iterations is None
    assert summary.failed_cases == 0


# --------------------------------------------------------------------- #
# Group: successful-but-unconverged cases
# --------------------------------------------------------------------- #
def test_unconverged_success_excluded_from_fastest_convergence_and_warns(tmp_path):
    excel = _build(tmp_path, [
        {"status": STATUS_DONE, "cl_cd": 10.0, "lift": 50.0, "drag": 5.0,
         "iterations": 100, "converged": "NO"},
        {"status": STATUS_DONE, "cl_cd": 8.0, "lift": 40.0, "drag": 5.0,
         "iterations": 450, "converged": "YES"},
    ])
    summary = analyze_study(excel, rows=[2, 3])
    assert summary.successful_cases == 2
    # The unconverged row's 100 iterations must NOT win "fastest convergence".
    assert summary.fastest_convergence_iterations == 450
    assert summary.fastest_convergence_row == 3
    warning = next(w for w in summary.warnings
                  if w.code == WarningCode.UNCONVERGED_SUCCESS)
    assert "1" in warning.message


# --------------------------------------------------------------------- #
# Group: deterministic tie-breaking — first row encountered wins
# --------------------------------------------------------------------- #
def test_tied_best_l_over_d_selects_the_first_row_deterministically(tmp_path):
    excel = _build(tmp_path, [
        {"status": STATUS_DONE, "cl_cd": 12.5, "lift": 50.0, "drag": 5.0,
         "iterations": 400, "converged": "YES"},
        {"status": STATUS_DONE, "cl_cd": 12.5, "lift": 50.0, "drag": 5.0,
         "iterations": 400, "converged": "YES"},
        {"status": STATUS_DONE, "cl_cd": 12.5, "lift": 50.0, "drag": 5.0,
         "iterations": 400, "converged": "YES"},
    ])
    # Run twice to prove the result isn't an accident of dict/set ordering.
    for _ in range(2):
        summary = analyze_study(excel, rows=[2, 3, 4])
        assert summary.best_l_over_d == 12.5
        assert summary.best_l_over_d_row == 2      # first row, never row 3 or 4
        assert summary.highest_lift_row == 2
        assert summary.lowest_drag_row == 2
        assert summary.fastest_convergence_row == 2


# --------------------------------------------------------------------- #
# Group: retries + case-failed are both stable, explicit rules
# --------------------------------------------------------------------- #
def test_retries_and_case_failed_are_independent_stable_rules(tmp_path):
    excel = _build(tmp_path, [{"status": "FAILED"}])
    summary = analyze_study(excel, rows=[2], retries=3)
    assert summary.retries == 3
    by_code = {w.code: w for w in summary.warnings}
    assert by_code[WarningCode.RETRIES_OCCURRED].message == \
        "3 retry attempt(s) were needed across this batch."
    assert by_code[WarningCode.CASE_FAILED].message == "1 case(s) failed."


def test_analyze_study_never_logs(tmp_path, caplog):
    """study_analytics must stay purely computational — logging is the
    orchestrator's responsibility, not this module's."""
    excel = _build(tmp_path, [{"status": "FAILED"}])
    with caplog.at_level("DEBUG"):
        analyze_study(excel, rows=[2], retries=1)
    assert caplog.records == []


# --------------------------------------------------------------------- #
# Group: orchestrator integration — wiring, lifecycle, and retry counting
# --------------------------------------------------------------------- #
class _FailsOnceThenSucceeds:
    """Fake SolverBackend: raises once (forcing a retry), then succeeds."""

    def __init__(self):
        self.calls = 0

    def run_case(self, exp, mesh_file, case_dir):
        self.calls += 1
        if self.calls == 1:
            raise DivergedError("transient failure (test fixture)")
        return CaseResult(cl=0.6, cd=0.05, lift_n=10.0, drag_n=1.0,
                          iterations=200, converged=True,
                          artifact_dir=str(case_dir))


def _cfg_and_excel(tmp_path, retries_per_case: int = 1):
    # build_template() seeds 8 example rows (4 AOA x 2 velocities); these
    # integration tests need exactly one row so retry/summary counts are
    # unambiguous, so clear the template down to a single PENDING row.
    xlsx = tmp_path / "e.xlsx"
    build_template(xlsx)
    wb = load_workbook(xlsx)
    ws = wb["Experiments"]
    _clear_rows(ws)
    h = _headers(ws)
    ws.cell(2, h["AOA_deg"], 5.0)
    ws.cell(2, h["Velocity_m_s"], 20.0)
    wb.save(xlsx)

    cfg_file = tmp_path / "c.yaml"
    cfg_file.write_text(f"""
fluent:
  aoa_method: "velocity_vector"
  wall_zones: ["wing"]
  reference: {{density: 1.225, area: 1.0}}
excel:
  file: "{xlsx.as_posix()}"
runtime:
  work_dir: "{(tmp_path / 'runs').as_posix()}"
  retries_per_case: {retries_per_case}
""")
    cfg = load_config(cfg_file)
    return cfg, ExcelManager(cfg.excel)


def test_orchestrator_populates_summary_and_counts_retries(tmp_path):
    cfg, excel = _cfg_and_excel(tmp_path, retries_per_case=1)
    orch = Orchestrator(cfg, excel, None, _FailsOnceThenSucceeds(), bus=EventBus())
    failures = orch.run(max_cases=1)

    assert failures == 0
    assert orch._retries_used == 1
    summary = orch._current_study_summary
    assert summary is not None
    assert summary.successful_cases == 1
    assert summary.retries == 1
    assert any(w.code == WarningCode.RETRIES_OCCURRED for w in summary.warnings)


def test_current_study_summary_and_retries_do_not_leak_between_runs(tmp_path):
    cfg, excel = _cfg_and_excel(tmp_path, retries_per_case=1)
    orch = Orchestrator(cfg, excel, None, _FailsOnceThenSucceeds(), bus=EventBus())
    orch.run(max_cases=1)
    assert orch._retries_used == 1   # sanity check on the first run

    # Nothing left to run (the one row is now DONE) — a second run() call
    # must reset both the retry counter and the summary, not carry the
    # previous study's numbers forward.
    failures = orch.run()
    assert failures == 0
    assert orch._retries_used == 0
    summary = orch._current_study_summary
    assert summary is not None
    assert summary.total_cases == 0
    assert summary.retries == 0
    assert summary.warnings and summary.warnings[0].code == WarningCode.EMPTY_STUDY
