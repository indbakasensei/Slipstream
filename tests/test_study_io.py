"""Universal CFD Platform, Phase 5 — StudyIO tests.

StudyIO is the template-driven boundary that maps a study's spreadsheet
representation to/from the runtime model: column resolution (honouring
ColumnMap overrides), row interpretation (blank/partial/unreadable rules),
Experiment construction, and export metadata. These tests lock in that
mapping, the workbook round-trip, and the byte-compatibility of import.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto.config import ColumnMap, ExcelConfig                # noqa: E402
from cfdauto.excel_manager import ExcelManager                    # noqa: E402
from cfdauto.experiment_definition import ExperimentDefinition    # noqa: E402
from cfdauto.study_io import StudyIO                              # noqa: E402
from tools.make_experiment_template import build_template         # noqa: E402


def _io(column_map: ColumnMap | None = None) -> StudyIO:
    return StudyIO.default(column_map or ColumnMap())


# --------------------------------------------------------------------- #
# Group: column resolution (template + ColumnMap overrides)
# --------------------------------------------------------------------- #
def test_default_column_headers_match_the_column_map_defaults():
    io = _io()
    assert io.input_parameter_names() == ["aoa", "velocity"]
    assert io.input_column_headers() == ["AOA_deg", "Velocity_m_s"]


def test_column_headers_honour_user_column_map_overrides():
    io = _io(ColumnMap(aoa="Alpha_deg", velocity="Vinf_m_s"))
    # The study parameter name is the ColumnMap attribute name, so a rename
    # flows through automatically.
    assert io.input_column_headers() == ["Alpha_deg", "Vinf_m_s"]


# --------------------------------------------------------------------- #
# Group: row interpretation (byte-identical skip rules)
# --------------------------------------------------------------------- #
def test_interpret_full_row_builds_a_generic_experiment():
    io = _io()
    exp, warn = io.interpret_row(2, {"aoa": 4.0, "velocity": 20.0},
                                 {"Flap": 15.0}, "PENDING")
    assert warn is None
    assert exp.aoa_deg == 4.0 and exp.velocity == 20.0
    assert exp.extra_wb_params == {"Flap": 15.0}
    assert exp.parameters["Flap"].source == "wbp"
    assert exp.status == "PENDING"


def test_interpret_blank_row_is_skipped_silently():
    exp, warn = _io().interpret_row(9, {"aoa": None, "velocity": None}, {}, "")
    assert exp is None and warn is None


def test_interpret_partial_row_is_skipped_with_a_warning():
    exp, warn = _io().interpret_row(4, {"aoa": 6.0, "velocity": None}, {}, "")
    assert exp is None
    assert warn is not None and "velocity" in warn


def test_interpret_unreadable_row_is_skipped_with_a_warning():
    exp, warn = _io().interpret_row(5, {"aoa": "oops", "velocity": 20.0}, {}, "")
    assert exp is None
    assert warn is not None and "unreadable" in warn


# --------------------------------------------------------------------- #
# Group: validation surface (delegates to ExperimentDefinition)
# --------------------------------------------------------------------- #
def test_validate_row_uses_template_bounds():
    io = _io()
    assert io.validate_row({"aoa": 5.0, "velocity": 20.0}) == []
    assert io.validate_row({"aoa": 200.0, "velocity": 20.0})   # above AOA max


# --------------------------------------------------------------------- #
# Group: export metadata
# --------------------------------------------------------------------- #
def test_export_headers_and_default_rows_come_from_the_template():
    io = _io()
    assert io.export_input_headers() == ["AOA_deg", "Velocity_m_s"]
    assert len(io.default_experiment_rows()) == 8


# --------------------------------------------------------------------- #
# Group: workbook round-trip through the real ExcelManager
# --------------------------------------------------------------------- #
def test_generated_workbook_round_trips_through_study_io(tmp_path):
    """Generate → read back: 8 experiments with the exact AOA/Velocity grid,
    proving the export and import boundaries agree."""
    path = tmp_path / "e.xlsx"
    build_template(path)
    exps = ExcelManager(ExcelConfig(file=str(path))).read_experiments()
    got = [(e.aoa_deg, e.velocity) for e in exps]
    assert got == [(0, 20), (0, 30), (4, 20), (4, 30),
                   (8, 20), (8, 30), (12, 20), (12, 30)]


def test_import_skips_blank_partial_and_unreadable_rows(tmp_path):
    """The reader's skip behaviour is unchanged end-to-end through StudyIO."""
    path = tmp_path / "e.xlsx"
    build_template(path)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    h = {c.value: i + 1 for i, c in enumerate(ws[1])}
    # row 10: blank spacer (leave empty); row 11: half-filled; row 12: junk
    ws.cell(11, h["AOA_deg"], 6.0)
    ws.cell(12, h["AOA_deg"], "junk"); ws.cell(12, h["Velocity_m_s"], 20.0)
    wb.save(path)
    exps = ExcelManager(ExcelConfig(file=str(path))).read_experiments()
    assert len(exps) == 8            # the 8 valid rows only; 10/11/12 skipped


def test_read_with_renamed_columns_via_column_map(tmp_path):
    """A user who renamed the input columns (sheet + ColumnMap) still reads."""
    path = tmp_path / "e.xlsx"
    build_template(path)
    wb = load_workbook(path)
    ws = wb["Experiments"]
    ws.cell(1, 1).value = "Alpha_deg"
    ws.cell(1, 2).value = "Vinf_m_s"
    wb.save(path)
    cm = ColumnMap(aoa="Alpha_deg", velocity="Vinf_m_s")
    exps = ExcelManager(ExcelConfig(file=str(path), columns=cm)).read_experiments()
    assert len(exps) == 8
    assert exps[0].aoa_deg == 0.0 and exps[0].velocity == 20.0
