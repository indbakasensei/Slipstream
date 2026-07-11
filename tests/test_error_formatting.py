"""Sprint 2 — behavioral tests for cfdauto.error_formatting.

format_error() is the single place that turns a raw exception into
something a user can act on. Before this module existed, every display
surface (CLI, GUI dialogs, doctor's table) printed `str(exc)` or
`f"{type(exc).__name__}: {exc}"` independently and inconsistently. These
tests protect two things: that each exception type gets the *specific*
explanation it should (not a generic one), and that the formatter can
never itself become a new source of crashes — it sits on every error path
in the application, so a bug in it would be worse than the error it was
explaining.
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cfdauto.config import load_config                           # noqa: E402
from cfdauto.error_formatting import format_error                # noqa: E402
from cfdauto.events import EventBus                               # noqa: E402
from cfdauto.excel_manager import ExcelManager                    # noqa: E402
from cfdauto.exceptions import (                                 # noqa: E402
    CaseError,
    CFDAutoError,
    ConfigError,
    DivergedError,
    ExcelWriteError,
    FluentError,
    FrameworkError,
    MeshNotFoundError,
    NotConvergedError,
    ResultExtractionError,
    WorkbenchError,
)
from cfdauto.orchestrator import Orchestrator                     # noqa: E402
from tools.make_experiment_template import build_template         # noqa: E402


# --------------------------------------------------------------------- #
# Group: exception-type-driven classification
#
# Regression Scenario: a case fails with a MeshNotFoundError vs. a
# DivergedError vs. an ExcelWriteError — three completely different
# problems (a Workbench export glob mismatch, a genuine solver blow-up, a
# locked spreadsheet) that all used to render as the same bare
# "ClassName: message" text. A user acting on the wrong guidance wastes
# real time (e.g. tweaking solver relaxation for what's actually a locked
# Excel file).
# Expected Behaviour: each specific exception type gets its own title and
# specific possible-reasons/suggested-action pair, driven by the
# exception's type alone — no message inspection required for these.
# Why this test exists: this is the core value proposition of the
# formatter; a title mix-up here is worse than the raw string it replaces.
# --------------------------------------------------------------------- #
def test_specific_exception_types_get_specific_distinct_titles():
    cases = [
        (MeshNotFoundError("no mesh"), "Mesh Not Found"),
        (DivergedError("CL non-finite"), "Solver Diverged"),
        (NotConvergedError("max iters hit"), "Solver Did Not Converge"),
        (ResultExtractionError("bad report def"), "Result Extraction Error"),
        (ExcelWriteError("locked"), "Excel Workbook Error"),
        (WorkbenchError("journal failed"), "Workbench Error"),
    ]
    seen_titles = set()
    for exc, expected_title in cases:
        fe = format_error(exc)
        assert fe.title == expected_title
        assert fe.summary == str(exc)
        assert fe.possible_reasons, f"{expected_title} must have reasons"
        assert fe.suggested_action, f"{expected_title} must have a suggested action"
        seen_titles.add(fe.title)
    assert len(seen_titles) == len(cases), "each type must get a distinct title"


def test_generic_case_and_application_and_unexpected_fallbacks():
    # A CaseError subclass with no dedicated profile falls back to the
    # generic CaseError-level explanation, not to "Unexpected Error".
    class SomeFutureCaseError(CaseError):
        pass

    assert format_error(SomeFutureCaseError("x")).title == "Case Error"
    assert format_error(CFDAutoError("x")).title == "Application Error"
    # A plain, non-framework exception (a real bug) is clearly labelled
    # "Unexpected" rather than being dressed up as a known failure mode.
    assert format_error(ValueError("boom")).title == "Unexpected Error"


# --------------------------------------------------------------------- #
# Group: message-refined classification (only for the broad types)
#
# Regression Scenario: ConfigError alone covers wildly different real
# problems — a typo'd YAML key vs. a missing baseline case vs. ANSYS not
# being installed at all — each needing different guidance (README's own
# Troubleshooting section treats them as separate FAQ entries). Likewise
# a FluentError from a license lockout needs completely different advice
# than a FluentError from a bad zone name.
# Expected Behaviour: for these specific, broad types only, the
# exception's own message is inspected to pick the right explanation;
# anything unmatched falls back to a still-useful generic message for
# that type (never blank, never crashes).
# Why this test exists: proves message-matching is used exactly where it
# was scoped to (broad types only) and that it actually distinguishes the
# scenarios README already documents as different.
# --------------------------------------------------------------------- #
def test_config_error_message_refinement():
    unknown = format_error(ConfigError("Unknown top-level config section(s): ['flunt']"))
    missing_ansys = format_error(ConfigError(
        "Cannot locate ANSYS: AWP_ROOT252 is not set, ansys.awp_root is empty"))
    missing_file = format_error(ConfigError("Baseline case file not found: C:/x.cas.h5"))
    bad_physics = format_error(ConfigError("Config validation failed:\n  - dimension must be 2 or 3"))
    fallback = format_error(ConfigError("some other problem"))

    assert "typo" in " ".join(unknown.possible_reasons).lower() or "renamed" in " ".join(unknown.possible_reasons).lower()
    assert "awp_root" in " ".join(missing_ansys.possible_reasons).lower() or "installed" in " ".join(missing_ansys.possible_reasons).lower()
    assert "moved" in " ".join(missing_file.possible_reasons).lower() or "renamed" in " ".join(missing_file.possible_reasons).lower()
    assert "inconsistent" in " ".join(bad_physics.possible_reasons).lower()
    assert fallback.possible_reasons                       # generic, but never empty
    # All five are still classified as the same title (Configuration Error).
    assert all(fe.title == "Configuration Error"
               for fe in (unknown, missing_ansys, missing_file, bad_physics, fallback))
    # But the guidance itself genuinely differs between scenarios.
    assert unknown.suggested_action != missing_ansys.suggested_action != missing_file.suggested_action


def test_fluent_error_distinguishes_license_lockout_from_generic():
    lockout = format_error(FluentError("Fluent failed to launch: RPC UNAVAILABLE"))
    generic = format_error(FluentError("Could not configure report definition 'cl-report'"))
    assert "license" in " ".join(lockout.possible_reasons).lower()
    assert lockout.suggested_action != generic.suggested_action


def test_framework_error_distinguishes_lock_from_generic():
    lock = format_error(FrameworkError("Another run (PID 1234) already owns runs/cfdauto.lock."))
    generic = format_error(FrameworkError("geometry mode requires a Workbench backend"))
    assert "lock" in " ".join(lock.possible_reasons).lower() or "run" in " ".join(lock.possible_reasons).lower()
    assert lock.suggested_action != generic.suggested_action


# --------------------------------------------------------------------- #
# Group: rendering contract
#
# Regression Scenario: the sprint spec requires specific section labels
# ("Possible Reasons", "Suggested Next Step", "Need more information?")
# and explicitly no separate "Category:" line since the title already
# carries that. A rendering regression here would silently drift the
# actual displayed text away from the agreed design.
# Expected Behaviour: render_text() uses exactly those labels, omits the
# "Need more information?" block when no location was supplied, and never
# prints a "Category:" line. render_compact() is always a single line.
# Why this test exists: locks down the approved output contract.
# --------------------------------------------------------------------- #
def test_render_text_labels_and_no_category_line():
    text = format_error(DivergedError("CL non-finite at iteration 240")).render_text()
    assert "Possible Reasons:" in text
    assert "Suggested Next Step:" in text
    assert "Category:" not in text
    assert "Need more information?" not in text     # no location supplied


def test_render_text_includes_location_only_when_supplied(tmp_path):
    case_dir = tmp_path / "r001_aoa0_v20"
    case_dir.mkdir()
    with_location = format_error(DivergedError("x"), case_dir=case_dir).render_text()
    assert "Need more information?" in with_location
    assert str(case_dir / "case.log") in with_location

    without_location = format_error(DivergedError("x")).render_text()
    assert "Need more information?" not in without_location

    log_path = tmp_path / "logs" / "cfdauto.log"
    with_log_path = format_error(DivergedError("x"), log_path=log_path).render_text()
    assert str(log_path) in with_log_path


def test_render_compact_is_single_line():
    compact = format_error(ExcelWriteError("Could not save x.xlsx")).render_compact()
    assert "\n" not in compact
    assert compact.startswith("Excel Workbook Error:")


# --------------------------------------------------------------------- #
# Group: the formatter itself must never raise
#
# Regression Scenario: format_error() is called from exception handlers
# across the CLI, GUI, orchestrator, and doctor — the worst possible place
# for a new bug to live, since a crash there would replace an
# understandable error with an unhandled one, or (in the GUI) potentially
# crash the error-reporting dialog itself.
# Expected Behaviour: even a maximally hostile exception (whose own
# __str__ raises) still produces a usable FormattedError instead of
# propagating.
# Why this test exists: this is the safety property the whole design
# leans on to be trustworthy to wire into every error path.
# --------------------------------------------------------------------- #
def test_format_error_never_raises_even_for_a_hostile_exception():
    class HostileError(RuntimeError):
        def __str__(self):
            raise RuntimeError("str() itself is broken")

    fe = format_error(HostileError())
    assert fe.title
    assert isinstance(fe.render_text(), str)
    assert isinstance(fe.render_compact(), str)


# --------------------------------------------------------------------- #
# Group: orchestrator integration — wiring, not just the formatter itself
#
# Regression Scenario: Sprint 2 requires the formatted explanation to
# reach the one place an engineer actually opens after a FAILED row: the
# per-case log file. It also requires that nothing about the *existing*
# workflow (the Excel Error column, result.json) changed shape in the
# process — this sprint is explicitly display-only.
# Expected Behaviour: after a case fails permanently, case_dir/case.log
# contains the structured "Possible Reasons"/"Suggested Next Step" block,
# while the Excel Error cell keeps its original terse
# "ExceptionType: message" format, unchanged from before this sprint.
# Why this test exists: a unit test on format_error() alone can't catch a
# wiring mistake (e.g. the orchestrator forgetting to pass case_dir, or
# accidentally changing what gets written to Excel).
# --------------------------------------------------------------------- #
class _AlwaysDivergesFluent:
    """Fake SolverBackend — deterministically fails every case."""

    def run_case(self, exp, mesh_file, case_dir):
        raise DivergedError("CL became non-finite at iteration 42 (test fixture)")


def test_orchestrator_writes_formatted_block_to_case_log_on_failure(tmp_path):
    xlsx = tmp_path / "e.xlsx"
    build_template(xlsx)
    cfg_file = tmp_path / "c.yaml"
    cfg_file.write_text(f"""
fluent:
  aoa_method: "velocity_vector"
  wall_zones: ["wing"]
  reference: {{density: 1.225, area: 1.0}}
excel:
  file: "{xlsx.as_posix()}"
runtime:
  work_dir: "{(tmp_path / 'runs').as_posix()}"
  retries_per_case: 0
""")
    cfg = load_config(cfg_file)
    excel = ExcelManager(cfg.excel)
    orch = Orchestrator(cfg, excel, None, _AlwaysDivergesFluent(), bus=EventBus())
    failures = orch.run(max_cases=1)
    assert failures == 1

    case_dirs = list((cfg.work_dir() / "cases").iterdir())
    assert len(case_dirs) == 1
    case_log = (case_dirs[0] / "case.log").read_text(encoding="utf-8")
    assert "Solver Diverged" in case_log
    assert "Possible Reasons:" in case_log
    assert "Suggested Next Step:" in case_log
    assert str(case_dirs[0] / "case.log") in case_log   # location points at itself

    # The Excel Error column keeps its pre-Sprint-2 terse format untouched.
    from openpyxl import load_workbook
    ws = load_workbook(xlsx)["Experiments"]
    headers = {c.value: i + 1 for i, c in enumerate(ws[1])}
    error_cell = str(ws.cell(2, headers["Error"]).value)
    assert error_cell == "DivergedError: CL became non-finite at iteration 42 (test fixture)"
    assert "Possible Reasons" not in error_cell         # Excel stays terse, unchanged
