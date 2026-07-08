#!/usr/bin/env python3
"""Generate the experiments.xlsx schedule template.

The workbook this produces is the *entire user interface* of the framework:
one row per experiment, inputs on the left, results (written by the code as
plain values, never formulas) on the right. Only styling and headers are
created here — the automation preserves the user's formatting when it writes
results back in place.

Run directly (``python tools/make_experiment_template.py [path]``) or via
``python main.py init-template [path]``.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Keep in sync with cfdauto.config.ColumnMap defaults.
INPUT_HEADERS = ["AOA_deg", "Velocity_m_s"]
OUTPUT_HEADERS = ["Status", "CL", "CD", "CL/CD", "Lift_N", "Drag_N", "FL/FD",
                  "Iterations", "Converged", "Error", "Started", "Finished",
                  "Duration_min", "CaseDir"]
WIDTHS = {"AOA_deg": 10, "Velocity_m_s": 13, "Status": 10, "CL": 10, "CD": 10,
          "CL/CD": 9, "Lift_N": 11, "Drag_N": 11, "FL/FD": 9, "Iterations": 11,
          "Converged": 11, "Error": 40, "Started": 19, "Finished": 19,
          "Duration_min": 13, "CaseDir": 34}

# A small, sensible starting matrix the user can edit/extend freely.
EXAMPLE_AOA = [0, 4, 8, 12]
EXAMPLE_VEL = [20, 30]

FONT = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark steel blue
INPUT_FILL = PatternFill("solid", fgColor="DDEBF7")    # light blue: "yours"


def build_template(path: Path) -> Path:
    wb = Workbook()

    # ------------------------------------------------------------- sheet 1
    ws = wb.active
    ws.title = "Experiments"
    headers = INPUT_HEADERS + OUTPUT_HEADERS
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = WIDTHS.get(name, 12)

    row = 2
    for aoa in EXAMPLE_AOA:
        for vel in EXAMPLE_VEL:
            a = ws.cell(row=row, column=1, value=float(aoa))
            v = ws.cell(row=row, column=2, value=float(vel))
            for c in (a, v):
                c.font = Font(name=FONT, size=11)
                c.fill = INPUT_FILL
                c.number_format = "0.0"
            row += 1

    ws.freeze_panes = "A2"          # keep the header visible while scrolling

    # ------------------------------------------------------------- sheet 2
    doc = wb.create_sheet("ReadMe")
    doc.column_dimensions["A"].width = 16
    doc.column_dimensions["B"].width = 96
    lines = [
        ("cfdauto experiment schedule", ""),
        ("", ""),
        ("How to use", "One experiment per row. Fill AOA_deg and Velocity_m_s, "
         "leave Status empty, save, then run:  python main.py run"),
        ("Resume", "The Status column IS the run state. Re-running the command "
         "only executes rows that are not DONE/SKIP. Rows left RUNNING by a "
         "crash are re-queued automatically."),
        ("", ""),
        ("AOA_deg", "Angle of attack in degrees (input)."),
        ("Velocity_m_s", "Freestream / inlet velocity magnitude in m/s (input)."),
        ("Status", "empty/PENDING = will run · RUNNING = in progress · DONE · "
         "FAILED (reason in Error) · SKIP = type this to exclude a row."),
        ("CL / CD", "Lift / drag coefficients from Fluent report definitions, "
         "normalised by the per-case reference values (rho, V_case, A)."),
        ("CL/CD", "Aerodynamic efficiency (computed value, not a formula)."),
        ("Lift_N / Drag_N", "Wind-axis force components in Newtons from Fluent "
         "'force' report definitions on the configured wall zones."),
        ("FL/FD", "Force ratio. With one set of reference values this equals "
         "CL/CD by definition — both are written because the sheet asks for both."),
        ("Iterations", "Iterations actually run for this case."),
        ("Converged", "YES when CL and CD were flat within tolerance over the "
         "convergence window; NO if max_iterations was reached first."),
        ("Error", "First 500 characters of the failure reason, if any."),
        ("Started/Finished/Duration_min", "Wall-clock bookkeeping per case."),
        ("CaseDir", "Per-case artifact folder: transcript, CL/CD history, "
         "wb journal + status, result.json, case log."),
        ("", ""),
        ("Extra design variables", "Add a column named  WBP:<WorkbenchParam>  "
         "(e.g. WBP:P3 or WBP:FlapAngle). Its value is pushed to that Workbench "
         "parameter before meshing — no code changes needed. Rows with the same "
         "AOA + WBP values share one mesh automatically."),
        ("Renaming columns", "Headers are configurable in config.yaml under "
         "excel.columns if you prefer different names/units."),
    ]
    for r, (k, v) in enumerate(lines, start=1):
        a = doc.cell(row=r, column=1, value=k)
        b = doc.cell(row=r, column=2, value=v)
        a.font = Font(name=FONT, bold=bool(k and not v) or r == 1, size=11)
        b.font = Font(name=FONT, size=11)
        b.alignment = Alignment(wrap_text=True, vertical="top")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("experiments.xlsx")
    print(f"Written: {build_template(target)}")
